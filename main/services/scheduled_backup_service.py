"""
Scheduled Backup Service
Handles automated PostgreSQL database dumps and Excel data exports.
Backups are stored in BASE_DIR/backups/ with automatic cleanup (keeps last 14).
"""

import os
import threading
import time
import subprocess
import logging
from datetime import datetime, timedelta
from pathlib import Path

from django.conf import settings

logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
BACKUP_DIR = Path(getattr(settings, 'BASE_DIR', '/var/inspection-system')) / 'backups'
DB_BACKUP_DIR = BACKUP_DIR / 'db'
EXCEL_BACKUP_DIR = BACKUP_DIR / 'excel'
KEEP_LAST = 14          # number of each type to keep
SCHEDULE_HOURS = 24     # run every N hours


# ── Internal state ─────────────────────────────────────────────────────────────
class _BackupServiceState:
    def __init__(self):
        self.is_running = False
        self._thread = None
        self._stop_event = threading.Event()
        self.last_backup_time = None
        self.last_backup_status = None
        self.last_backup_message = ''
        self.next_backup_time = None

    # ── Lifecycle ──────────────────────────────────────────────────────────────
    def start(self):
        if self.is_running:
            return True, 'Backup service already running'
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._loop, daemon=True, name='backup-service')
        self._thread.start()
        self.is_running = True
        self.next_backup_time = datetime.now() + timedelta(hours=SCHEDULE_HOURS)
        logger.info('Backup service started')
        return True, 'Backup service started'

    def stop(self):
        if not self.is_running:
            return True, 'Backup service not running'
        self._stop_event.set()
        self.is_running = False
        self.next_backup_time = None
        logger.info('Backup service stopped')
        return True, 'Backup service stopped'

    def run_now(self):
        """Run a backup immediately (in calling thread — use for manual triggers)."""
        return _run_backup()

    # ── Background loop ────────────────────────────────────────────────────────
    def _loop(self):
        # Run once immediately on start
        success, msg = _run_backup()
        self.last_backup_time = datetime.now()
        self.last_backup_status = success
        self.last_backup_message = msg
        self.next_backup_time = datetime.now() + timedelta(hours=SCHEDULE_HOURS)

        while not self._stop_event.is_set():
            sleep_seconds = max(0, (self.next_backup_time - datetime.now()).total_seconds())
            self._stop_event.wait(timeout=sleep_seconds)
            if self._stop_event.is_set():
                break
            success, msg = _run_backup()
            self.last_backup_time = datetime.now()
            self.last_backup_status = success
            self.last_backup_message = msg
            self.next_backup_time = datetime.now() + timedelta(hours=SCHEDULE_HOURS)

    def get_status(self):
        return {
            'is_running': self.is_running,
            'status': 'running' if self.is_running else 'stopped',
            'last_backup_time': self.last_backup_time.isoformat() if self.last_backup_time else None,
            'last_backup_status': self.last_backup_status,
            'last_backup_message': self.last_backup_message,
            'next_backup_time': self.next_backup_time.isoformat() if self.next_backup_time else None,
            'backup_dir': str(BACKUP_DIR),
            'db_backups': _list_backups(DB_BACKUP_DIR, '*.dump'),
            'excel_backups': _list_backups(EXCEL_BACKUP_DIR, '*.xlsx'),
        }


_service = _BackupServiceState()


# ── Core backup logic ──────────────────────────────────────────────────────────
def _run_backup():
    """Run both a pg_dump and an Excel export. Returns (success, message)."""
    _ensure_dirs()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    errors = []

    # 1. PostgreSQL dump
    db_ok, db_msg = _pg_dump(timestamp)
    if not db_ok:
        errors.append(f'DB: {db_msg}')
    else:
        _cleanup(DB_BACKUP_DIR, '*.dump', KEEP_LAST)

    # 2. Excel export
    xl_ok, xl_msg = _excel_export(timestamp)
    if not xl_ok:
        errors.append(f'Excel: {xl_msg}')
    else:
        _cleanup(EXCEL_BACKUP_DIR, '*.xlsx', KEEP_LAST)

    if errors:
        msg = ' | '.join(errors)
        logger.error('Backup errors: %s', msg)
        return False, msg

    msg = f'Backup completed: {db_msg} | {xl_msg}'
    logger.info(msg)
    return True, msg


def _pg_dump(timestamp):
    db = settings.DATABASES['default']
    out_file = DB_BACKUP_DIR / f'backup_{timestamp}.dump'
    env = os.environ.copy()
    env['PGPASSWORD'] = db.get('PASSWORD', '')
    cmd = [
        'pg_dump',
        '-h', db.get('HOST', 'localhost'),
        '-p', str(db.get('PORT', 5432)),
        '-U', db.get('USER', 'postgres'),
        '-F', 'c',          # custom compressed format
        '-f', str(out_file),
        db.get('NAME', ''),
    ]
    try:
        result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=300)
        if result.returncode != 0:
            return False, f'pg_dump failed: {result.stderr.strip()}'
        size_kb = out_file.stat().st_size // 1024
        return True, f'DB dump saved ({size_kb} KB): {out_file.name}'
    except FileNotFoundError:
        return False, 'pg_dump not found — ensure postgresql-client is installed'
    except subprocess.TimeoutExpired:
        return False, 'pg_dump timed out after 5 minutes'
    except Exception as e:
        return False, str(e)


def _excel_export(timestamp):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.contrib.auth import get_user_model
        from ..models import (
            FoodSafetyAgencyInspection, Client, InspectorMapping,
            InspectorTarget, InspectorSalary, QuarterlyTarget,
            SystemLog,
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='007890')

        def _write_sheet(ws, headers, rows):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for row in rows:
                ws.append(row)
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        User = get_user_model()

        # ── Inspections ──────────────────────────────────────────────────────
        ws = wb.create_sheet('Inspections')
        qs = FoodSafetyAgencyInspection.objects.select_related('client').order_by('-date_of_inspection')[:5000]
        _write_sheet(ws,
            ['ID', 'Date', 'Inspector', 'Client', 'Commodity', 'Approved', 'Sent', 'Sent Date', 'COA Uploaded'],
            [
                [
                    i.id,
                    str(i.date_of_inspection or ''),
                    i.inspector_name or '',
                    i.client_name or '',
                    i.commodity or '',
                    i.approved_status or '',
                    'Yes' if i.is_sent else 'No',
                    str(i.sent_date or ''),
                    'Yes' if i.coa_uploaded_date else 'No',
                ]
                for i in qs
            ]
        )

        # ── Clients ──────────────────────────────────────────────────────────
        ws = wb.create_sheet('Clients')
        _write_sheet(ws,
            ['Client ID', 'Name', 'Account Code', 'Town', 'Corporate Group', 'Facility Type', 'Email'],
            [
                [c.client_id, c.eclick_name or '', c.internal_account_code or '',
                 c.town or '', c.corporate_group or '', c.facility_type or '',
                 c.representative_email or '']
                for c in Client.objects.all().order_by('eclick_name')
            ]
        )

        # ── Users ────────────────────────────────────────────────────────────
        ws = wb.create_sheet('Users')
        _write_sheet(ws,
            ['Username', 'First Name', 'Last Name', 'Email', 'Role', 'Active'],
            [
                [u.username, u.first_name, u.last_name, u.email, u.role or '', 'Yes' if u.is_active else 'No']
                for u in User.objects.all().order_by('username')
            ]
        )

        # ── Inspector Salaries ────────────────────────────────────────────────
        ws = wb.create_sheet('Inspector Salaries')
        _write_sheet(ws,
            ['Inspector Name', 'Monthly Salary', 'Employee Number', 'Updated At'],
            [
                [s.inspector_name, float(s.monthly_salary), s.employee_number or '', str(s.updated_at or '')]
                for s in InspectorSalary.objects.all().order_by('inspector_name')
            ]
        )

        # ── Inspector Targets ─────────────────────────────────────────────────
        ws = wb.create_sheet('Inspector Targets')
        _write_sheet(ws,
            ['Inspector', 'Eggs', 'Poultry', 'Raw', 'PMP', 'Raw Samples', 'PMP Samples', 'Total Samples'],
            [
                [t.inspector_name, t.eggs or 0, t.poultry or 0, t.raw or 0, t.pmp or 0,
                 t.raw_samples or 0, t.pmp_samples or 0, t.total_samples or 0]
                for t in InspectorTarget.objects.all().order_by('inspector_name')
            ]
        )

        # ── Quarterly Targets ─────────────────────────────────────────────────
        ws = wb.create_sheet('Quarterly Targets')
        _write_sheet(ws,
            ['Inspector', 'Year', 'Quarter', 'Monthly Salary', 'Quarterly Revenue Target', 'Monthly Vehicle Cost'],
            [
                [q.inspector_name, q.year, q.quarter,
                 float(q.monthly_salary or 0), float(q.quarterly_revenue_target or 0),
                 float(q.monthly_vehicle_cost or 0)]
                for q in QuarterlyTarget.objects.all().order_by('inspector_name', 'year', 'quarter')
            ]
        )

        # ── Recent System Logs ─────────────────────────────────────────────────
        ws = wb.create_sheet('System Logs (Recent)')
        _write_sheet(ws,
            ['Timestamp', 'User', 'Action', 'Page', 'Object Type', 'Description'],
            [
                [str(l.timestamp), l.user.username if l.user else '', l.action or '',
                 l.page or '', l.object_type or '', (l.description or '')[:200]]
                for l in SystemLog.objects.select_related('user').order_by('-timestamp')[:2000]
            ]
        )

        out_file = EXCEL_BACKUP_DIR / f'backup_{timestamp}.xlsx'
        wb.save(str(out_file))
        size_kb = out_file.stat().st_size // 1024
        return True, f'Excel saved ({size_kb} KB): {out_file.name}'

    except Exception as e:
        logger.exception('Excel export failed')
        return False, str(e)


# ── Helpers ────────────────────────────────────────────────────────────────────
def _ensure_dirs():
    DB_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def _list_backups(directory, pattern):
    if not directory.exists():
        return []
    files = sorted(directory.glob(pattern), key=lambda f: f.stat().st_mtime, reverse=True)
    return [
        {
            'name': f.name,
            'size_kb': f.stat().st_size // 1024,
            'created': datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
        }
        for f in files
    ]


def _cleanup(directory, pattern, keep):
    files = sorted(directory.glob(pattern), key=lambda f: f.stat().st_mtime, reverse=True)
    for old_file in files[keep:]:
        try:
            old_file.unlink()
            logger.info('Deleted old backup: %s', old_file.name)
        except Exception:
            pass


# ── Public API (called by views) ───────────────────────────────────────────────
def get_scheduled_backup_service_status():
    return _service.get_status()


def start_scheduled_backup_service():
    return _service.start()


def stop_scheduled_backup_service():
    return _service.stop()


def run_manual_backup():
    return _service.run_now()
