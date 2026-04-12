from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt as _csrf_exempt
from django.db import transaction
from django.utils.dateparse import parse_date
from ..models import Shipment, Client
from .utils import apply_filters, clear_messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import datetime

# ---------------------------------------------------------------------------
#  IP Geolocation caches (module-level — persist for server lifetime)
# ---------------------------------------------------------------------------
_IP_GEO_CACHE: dict = {}       # ip -> full geo dict
_LATLON_CACHE: dict = {}       # "lat_r,lon_r" -> nominatim detail dict

_PRIVATE_PREFIXES = (
    '127.', '10.', '::1',
    '172.16.', '172.17.', '172.18.', '172.19.', '172.20.', '172.21.',
    '172.22.', '172.23.', '172.24.', '172.25.', '172.26.', '172.27.',
    '172.28.', '172.29.', '172.30.', '172.31.',
    '192.168.',
)
_GEO_EMPTY = {'suburb': '', 'city': '', 'state': '', 'country': '',
              'country_code': '', 'display': ''}

def _is_private_ip(ip: str) -> bool:
    return any(ip.startswith(p) for p in _PRIVATE_PREFIXES)


def _nominatim_reverse(lat: float, lon: float, fallback: dict) -> dict:
    """Reverse-geocode a lat/lon using Nominatim. Returns suburb-level detail."""
    import requests as _req
    try:
        r = _req.get(
            'https://nominatim.openstreetmap.org/reverse',
            params={'lat': lat, 'lon': lon, 'format': 'json',
                    'zoom': 16, 'addressdetails': 1},
            headers={'User-Agent': 'FoodSafetyAgencyMgmt/1.0'},
            timeout=7,
        )
        if r.status_code != 200:
            return fallback
        d = r.json()
        addr = d.get('address', {})
        suburb = (addr.get('suburb') or addr.get('neighbourhood') or
                  addr.get('residential') or addr.get('hamlet') or
                  addr.get('quarter') or '')
        city = (addr.get('city') or addr.get('town') or
                addr.get('village') or addr.get('municipality') or
                fallback.get('city', ''))
        state   = addr.get('state', '')         or fallback.get('state', '')
        country = addr.get('country', '')       or fallback.get('country', '')
        cc      = addr.get('country_code', '').upper() or fallback.get('country_code', '')
        parts = []
        if suburb:                parts.append(suburb)
        if city and city != suburb: parts.append(city)
        if state:                 parts.append(state)
        if country:               parts.append(country)
        return {
            'suburb': suburb, 'city': city, 'state': state,
            'country': country, 'country_code': cc,
            'display': ', '.join(parts) if parts else fallback.get('display', ''),
        }
    except Exception:
        return fallback


def _resolve_ips_to_locations(ip_list: list) -> dict:
    """
    Two-step resolution:
      1. ip-api.com batch  → city / state / country / lat / lon  (fast, one HTTP call)
      2. Nominatim reverse → suburb detail using the lat/lon      (1 req per new IP)
    Both levels are cached in module-level dicts.
    """
    import requests as _req
    import time as _time

    # ── Step 1: ip-api batch for IPs not yet cached ──────────────────────────
    to_resolve = [
        ip for ip in ip_list
        if ip and ip not in _IP_GEO_CACHE and not _is_private_ip(ip)
    ]
    needs_nominatim = []   # (ip, lat, lon, fallback_dict)

    if to_resolve:
        try:
            payload = [
                {'query': ip,
                 'fields': 'status,country,countryCode,regionName,city,lat,lon,query'}
                for ip in to_resolve[:100]
            ]
            resp = _req.post('http://ip-api.com/batch', json=payload, timeout=8)
            if resp.status_code == 200:
                for item in resp.json():
                    ip  = item.get('query', '')
                    if item.get('status') == 'success':
                        city  = item.get('city', '')
                        state = item.get('regionName', '')
                        cntry = item.get('country', '')
                        cc    = item.get('countryCode', '')
                        lat   = item.get('lat')
                        lon   = item.get('lon')
                        seen  = [p for p in [city, state, cntry] if p]
                        # deduplicate
                        deduped = []
                        for p in seen:
                            if p not in deduped:
                                deduped.append(p)
                        fallback = {
                            'suburb': '', 'city': city, 'state': state,
                            'country': cntry, 'country_code': cc,
                            'display': ', '.join(deduped),
                        }
                        _IP_GEO_CACHE[ip] = fallback
                        if lat and lon:
                            needs_nominatim.append((ip, lat, lon, fallback))
                    else:
                        _IP_GEO_CACHE[ip] = _GEO_EMPTY.copy()
        except Exception:
            pass

    # ── Step 2: Nominatim reverse-geocode for suburb detail ──────────────────
    # Rate-limit: 1 req / sec (Nominatim fair-use policy).
    # Cap at 8 new IPs per API call so the request stays under ~10 s.
    processed = 0
    for ip, lat, lon, fallback in needs_nominatim:
        cache_key = f"{round(lat, 2)},{round(lon, 2)}"   # ~1 km cell
        if cache_key in _LATLON_CACHE:
            _IP_GEO_CACHE[ip] = _LATLON_CACHE[cache_key]
            continue
        if processed >= 8:
            break
        detail = _nominatim_reverse(lat, lon, fallback)
        _LATLON_CACHE[cache_key] = detail
        _IP_GEO_CACHE[ip] = detail
        processed += 1
        if processed < len(needs_nominatim):
            _time.sleep(1.1)   # Nominatim: ≤ 1 req/s

    # ── Build result map ──────────────────────────────────────────────────────
    result = {}
    for ip in ip_list:
        if not ip or _is_private_ip(ip):
            result[ip] = _GEO_EMPTY.copy()
        else:
            result[ip] = _IP_GEO_CACHE.get(ip, _GEO_EMPTY.copy())
    return result
import os
import threading
import time
import shutil
from pathlib import Path
import re
import sys

# Add the project root to Python path to import the config
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

try:
    from eclick_mysql_config import (
        MYSQL_CONFIG, 
        SQLSERVER_CONFIG, 
        SQLSERVER_CONNECTION_STRING,
        FSA_INSPECTION_QUERY
    )
except ImportError:
    # Fallback configurations if import fails
    MYSQL_CONFIG = {
        'host': '77.37.121.135',
        'user': 'admin',
        'password': 'mk7z@Geg123',
        'database': 'E-click-Project-management',
        'port': 3306,
        'charset': 'utf8mb4',
        'autocommit': True,
        'connect_timeout': 60,
        'read_timeout': 60,
        'write_timeout': 60,
        'use_unicode': True,
    }
    
    SQLSERVER_CONFIG = {
        'ENGINE': 'mssql',
        'NAME': 'AFS',
        'USER': 'FSAUser2',
        'PASSWORD': 'password',
        'HOST': '102.67.140.12',
        'PORT': '1053',
        'OPTIONS': {
            'driver': 'ODBC Driver 17 for SQL Server',
            'trusted_connection': 'no',
        },
    }
    
    SQLSERVER_CONNECTION_STRING = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=102.67.140.12,1053;'
        'DATABASE=AFS;'
        'UID=FSAUser2;'
        'PWD=password;'
        'TrustServerCertificate=yes;'
        'Encrypt=yes;'
    )
    
    # Fallback FSA query (comprehensive inspection data) - FROM OCTOBER 2025 ONWARDS
    # IMPORTANT: Removed DISTINCT so that one inspection with multiple products = multiple rows (one per product)
    # IMPORTANT: Now includes InternalAccountNumber to match with Google Sheets client names
    # IMPORTANT: Pulls inspections from 2025-10-01 onwards (no end date limit)
    # IMPORTANT: Changed GPS JOIN to OUTER APPLY to prevent duplicate inspections
    #            - Before: One inspection with 9 GPS records = 9 duplicate inspection rows
    #            - After: One inspection with 9 GPS records = 1 inspection row (using first GPS record)
    FSA_INSPECTION_QUERY = '''
        SELECT 'POULTRY' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName as ProductName FROM AFS.dbo.PoultryLabelInspectionChecklistRecords OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryLabelChecklistInspectionRecordId = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ClientId where AFS.dbo.[PoultryLabelInspectionChecklistRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName IS NOT NULL AND [AFS].[dbo].[PoultryLabelInspectionChecklistRecords].ProductName != ''
        UNION ALL
        SELECT 'EGGS' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForInspection as IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, NULL AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PoultryEggInspectionRecords].Id as Id, clt.Name as Client, clt.InternalAccountNumber as InternalAccountNumber, [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer as ProductName FROM [AFS].[dbo].[PoultryEggInspectionRecords] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PoultryEggInspectionRecordId = [AFS].[dbo].[PoultryEggInspectionRecords].Id ORDER BY Id) gps join AFS.dbo.Clients clt on clt.Id = [AFS].[dbo].[PoultryEggInspectionRecords].ClientId where AFS.dbo.[PoultryEggInspectionRecords].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer IS NOT NULL AND [AFS].[dbo].[PoultryEggInspectionRecords].EggProducer != ''
        UNION ALL
        SELECT 'RAW' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.RawRMPInspectionLabSampleLinks WHERE RawRMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[RawRMPInspectionRecordTypes].Id as Id, COALESCE(prod.NewClientName, clt.Name) as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.NewProductItemDetails as ProductName FROM [AFS].[dbo].[RawRMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE RawRMPInspectionRecordId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.RawRMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[RawRMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[RawRMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND prod.NewProductItemDetails IS NOT NULL AND prod.NewProductItemDetails != ''
        UNION ALL
        SELECT 'PMP' as Commodity, DateOfInspection, StartOfInspection, EndOfInspection, InspectionLocationTypeID, IsDirectionPresentForthisInspection, InspectorId, gps.Latitude AS Latitude, gps.Longitude AS Longitude, CASE WHEN EXISTS (SELECT 1 FROM AFS.dbo.PMPInspectionLabSampleLinks WHERE PMPInspectionLabSampleLinks.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id) THEN 1 ELSE 0 END AS IsSampleTaken, NULL AS InspectionTravelDistanceKm, [AFS].[dbo].[PMPInspectionRecordTypes].Id as Id, COALESCE(prod.NewClientName, clt.Name) as Client, clt.InternalAccountNumber as InternalAccountNumber, prod.PMPItemDetails as ProductName FROM [AFS].[dbo].[PMPInspectionRecordTypes] OUTER APPLY (SELECT TOP 1 Latitude, Longitude FROM AFS.dbo.GPSInspectionLocationRecords WHERE PMPInspectionRecordId = [AFS].[dbo].[PMPInspectionRecordTypes].Id ORDER BY Id) gps JOIN AFS.dbo.PMPInspectedProductRecordTypes prod on prod.InspectionId = [AFS].[dbo].[PMPInspectionRecordTypes].Id join AFS.dbo.Clients clt on clt.Id = prod.ClientId where AFS.dbo.[PMPInspectionRecordTypes].IsActive = 1 AND DateOfInspection >= '2025-10-01' AND prod.PMPItemDetails IS NOT NULL AND prod.PMPItemDetails != ''
    '''

# Mapping of InspectorId to human-readable name (module-level so all views can use it)
INSPECTOR_NAME_MAP = {
    132: 'PAKI OLIFANT', 142: 'MARIANA DU TOIT', 143: 'MOKGADI SELONE', 140: 'AGREEMENT MOSIA',
    68: 'BEN VISAGIE', 144: 'VHAHANGWELE RALULIMI', 97: 'NAKISANI BALOYI', 133: 'CALVIN CLAASSENS',
    118: 'NEO NOE', 131: 'EDITH SELEPE', 124: 'SANDISIWE DLISANI', 85: 'JOEL MHANGWA',
    86: 'ASISIPHO LANDE', 125: 'MARIUS CARSTENS', 145: 'WILSON MAIFO', 82: 'JOE ROSENBLATT',
    95: 'EVANS NKWINIKA', 93: 'VICTOR MATHEBULA', 77: 'DELAREY RIBBENS', 78: 'DEWALD KORF',
    103: 'GERRIT PEKEMA', 76: 'PETRUS POOL', 71: 'PALESA MPANA', 134: 'ARMAND VISAGIE',
    113: 'KATLEGO MOKHUA', 147: 'BRIAN XULU', 146: 'ELIAS THEKISO', 72: 'CHARMAINE NEL',
    148: 'COLLEN DLAMINI', 153: 'JOFRED STEYN', 164: 'HENNIE CHILOANE', 166: 'THATO SEKHOTHO',
    69: 'AFS DUMMY', 177: 'CINGA NGONGO', 91: 'MONTI RAMAANO', 79: 'WENDY CHAKA',
    174: 'LWANDILE MAQINA', 96: 'MCAULEY MUSUNDA', 92: 'THABO MAGWAZA', 149: 'THEMBA SHABANGU',
    175: 'PETRUS POOL', 70: 'THERESA DIOGO', 73: 'RAM RAMBURAN', 74: 'HEIN NEL',
    75: 'SIBONELO ZONDI', 80: 'NIPHO NGOMANE', 81: 'CHELESILE MOYO', 83: 'SIMON SWART',
    84: 'ANDREAS LETABA', 87: 'LOUIS VISAGIE', 88: 'RASSIE SMIT', 89: 'NICOLE BERGH',
    90: 'Spare1 Spare2', 94: 'FRANCOIS PRETORIUS', 98: 'PLACEHOLDER PLACEHOLDER', 99: 'TEST GMAIL',
    102: 'ALI MODIKOE', 106: 'ARMAND VISAGIE', 111: 'DUMMY TESTER', 112: 'TESTER TESTER',
    114: 'DUMMY PLACEHOLDER', 115: 'TESTER PERSON', 116: 'FINANCE TWO', 119: 'TESTER TESTER',
    122: 'Spare Spare', 123: 'LIZELLE BREEDT', 126: 'THAPELO MAPOTSE', 127: 'THAPELO MAPOTSE',
    141: 'IT IT', 150: 'CHRISNA POOL', 154: 'SEUN SEBOLAI', 160: 'LERATO MODIBA',
    163: 'DENNIS CELE', 171: 'MARNUS BADENHORST', 173: 'HLENGIWE GUMEDE', 178: 'NTABISENG MASEKO',
    179: 'KABELO MOKGALAKA', 183: 'MPHO MOTAUNG', 184: 'XOLA MPELUZA', 185: 'PERCY MALEKA',
    186: 'KUTLWANO KUNTWANE', 187: 'SIPHO NDAMASE', 188: 'GLADYS MANGANYE', 193: 'ANTHONY PENZES',
    194: 'PHUMZILE MASOMBUKA', 196: 'NELISA NTOYAPHI',
}

# =============================================================================
# SQL DATABASE VIEWS
# =============================================================================

@login_required
def remote_sqlserver_data_view(request):
    """View to display data from remote SQL Server database (FSA Server)."""
    sqlserver_data = []
    error_message = None
    
    try:
        import pymssql

        # Use pymssql - NO ODBC DRIVERS NEEDED!
        connection = pymssql.connect(
            server='102.67.140.12',
            port=1053,
            user='FSAUser2',
            password='password',
            database='AFS',
            timeout=30
        )
        cursor = connection.cursor()
        
        # Execute the FSA inspection query
        query = FSA_INSPECTION_QUERY
        
        cursor.execute(query)
        
        # Fetch all results and convert to list of dictionaries
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            row_dict = dict(zip(columns, row))
            inspector_id = row_dict.get('InspectorId')
            try:
                # Ensure numeric for mapping
                inspector_id_int = int(inspector_id) if inspector_id is not None else None
            except (TypeError, ValueError):
                inspector_id_int = None
            # Add human-readable inspector name and keep id
            row_dict['Inspector'] = INSPECTOR_NAME_MAP.get(inspector_id_int, 'Unknown')
            sqlserver_data.append(row_dict)
        
        # Close connection
        cursor.close()
        connection.close()
        
    except ImportError as e:
        error_message = f"SQL Server connector not installed. Please install pyodbc. Error: {str(e)}"
    except pyodbc.Error as e:
        error_message = f"SQL Server connection error: {str(e)}"
    except Exception as e:
        error_message = f"Unexpected error: {str(e)}"
        import traceback
        traceback.print_exc()
    
    context = {
        'mysql_data': sqlserver_data,  # Reuse the same template
        'error_message': error_message,
        'data_count': len(sqlserver_data) if sqlserver_data else 0,
        'data_source': 'Remote FSA SQL Server Database - Inspection Records'
    }
    
    return render(request, 'main/remote_data.html', context)


# =============================================================================
# EXPORT VIEWS
# =============================================================================

@login_required(login_url='login')
def export_shipments(request):
    """Export shipment data to different formats (Excel, CSV, PDF)."""
    clear_messages(request)
    
    # Get export format and other parameters
    export_format = request.GET.get('format', 'excel')
    client_id = request.GET.get('client')
    
    # Get shipments with filters if provided - Use optimized queryset
    shipments = Shipment.objects.select_related('client').all()
    shipments = apply_filters(request, shipments)
    
    # Get client name for filename
    client_name = "all_clients"
    if client_id:
        try:
            client = Client.objects.get(pk=client_id)
            client_name = client.name.replace(" ", "_").replace("/", "_")
        except Client.DoesNotExist:
            pass
    
    # Generate timestamp for filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Create filename based on client and date
    filename_base = f"claims_{client_name}_{timestamp}"
    

    
    # Export based on selected format
    if export_format == 'excel':
        response = export_to_excel(shipments, filename_base)
        return response
    elif export_format == 'csv':
        response = export_to_csv(shipments, filename_base)
        return response
    elif export_format == 'pdf':
        response = export_to_pdf(shipments, filename_base)
        return response
    else:
        messages.error(request, f"Unsupported export format: {export_format}")
        return redirect('shipment_list')


@login_required(login_url='login')
def export_shipments_excel(request):
    """Legacy function that redirects to the more flexible export_shipments function."""
    return export_shipments(request)


# =============================================================================
# EXPORT HELPER FUNCTIONS - MATCHING TABLE COLUMNS EXACTLY
# =============================================================================

def export_to_excel(shipments, filename_base):
    """Helper function to export data to Excel format - matches table columns exactly."""
    # Create a workbook and active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Shipments'
    
    # Define headers exactly as shown in the table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row_num, shipment in enumerate(shipments, 2):
        # Use the new client-specific shipment numbers
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields as icons/text
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status badges
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓ Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗ Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~ Partial'
        else:
            settlement_status = '-'
        
        status_display = ''
        if shipment.Status == 'OPEN':
            status_display = '● Open'
        elif shipment.Status == 'CLOSED':
            status_display = '✓ Closed'
        elif shipment.Status == 'PENDING':
            status_display = '⏳ Pending'
        elif shipment.Status == 'REJECTED':
            status_display = '✗ Rejected'
        elif shipment.Status == 'UNDER_REVIEW':
            status_display = '◐ Under Review'
        else:
            status_display = shipment.Status
        
        # Create row data matching table exactly
        row_data = [
            shipment.Claim_No,  # This will now be the new format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Add row to worksheet
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        worksheet.column_dimensions[column_letter].width = adjusted_width
    

    
    # Create response for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
    
    # Save workbook to response
    workbook.save(response)
    return response


def export_to_csv(shipments, filename_base):
    """Helper function to export data to CSV format - matches table columns exactly."""
    # Create a file-like buffer for response
    response_buffer = io.StringIO()
    
    # Create CSV writer
    response_writer = csv.writer(response_buffer)
    
    # Write header row exactly as shown in table
    headers = [
        'Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 
        'Intent', 'Intent Date', 'Formal', 'Formal Date', 'Value', 
        'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 'Savings',
        'Settlement', 'Exposure', 'Status', 'Closed', 'Actions'
    ]
    response_writer.writerow(headers)
    
    # Write data rows
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "Yes" if shipment.Intent_To_Claim == 'YES' else "No"
        formal_claim = "Yes" if shipment.Formal_Claim_Received == 'YES' else "No"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = 'Settled'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = 'Not Settled'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = 'Partial'
        else:
            settlement_status = '-'
        
        status_display = shipment.get_Status_display() if shipment.Status else 'Open'
        
        row = [
            shipment.Claim_No,  # New format: ClientName-X-YYYYMMDD
            shipment.Brand or '-',
            shipment.Claimant or '-',
            client_id,
            client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_display,
            closed_date,
            'Edit/Delete'  # Actions column placeholder
        ]
        
        # Write to response
        response_writer.writerow(row)
    

    
    # Create response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write(response_buffer.getvalue())
    
    return response


def export_to_pdf(shipments, filename_base):
    """Helper function to export data to PDF format - matches table columns exactly."""
    # Create a file-like buffer for the PDF data
    buffer = io.BytesIO()
    
    # Create the PDF object with landscape orientation for all columns
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        title=f"Claims Report - {filename_base}",
        topMargin=20,
        bottomMargin=20,
        leftMargin=20,
        rightMargin=20
    )
    
    # Create PDF path for backup
    pdf_backup_path = os.path.join(backup_dir, 'pdf', f"{filename_base}.pdf")
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Add title
    title = Paragraph(f"Claims Report - {datetime.datetime.now().strftime('%Y-%m-%d')}", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))  # Add spacing
    
    # Define table data exactly matching the web table
    data = [
        ['Shipment No', 'Brand', 'Claimant', 'Claim ID', 'Client Name', 'Intent', 'Intent Date', 
         'Formal', 'Formal Date', 'Value', 'ISCM Paid', 'Carrier Paid', 'Insurance', 'Branch', 
         'Savings', 'Settlement', 'Exposure', 'Status', 'Closed']
    ]
    
    # Add shipment data
    for shipment in shipments:
        client_id = shipment.client.client_id if shipment.client else 'N/A'
        client_name = shipment.client.name if shipment.client else 'Unknown'
        
        # Format dates
        intend_date = shipment.Intend_Claim_Date.strftime("%m/%d/%y") if shipment.Intend_Claim_Date else '-'
        formal_date = shipment.Formal_Claim_Date_Received.strftime("%m/%d/%y") if shipment.Formal_Claim_Date_Received else '-'
        closed_date = shipment.Closed_Date.strftime("%m/%d/%y") if shipment.Closed_Date else '-'
        
        # Format amounts
        claimed_amount = f"${shipment.Claimed_Amount:,.0f}" if shipment.Claimed_Amount else "$0"
        iscm_paid = f"${shipment.Amount_Paid_By_Awa:,.0f}" if shipment.Amount_Paid_By_Awa else "$0"
        carrier_paid = f"${shipment.Amount_Paid_By_Carrier:,.0f}" if shipment.Amount_Paid_By_Carrier else "$0"
        insurance_paid = f"${shipment.Amount_Paid_By_Insurance:,.0f}" if shipment.Amount_Paid_By_Insurance else "$0"
        total_savings = f"${shipment.Total_Savings:,.0f}" if shipment.Total_Savings else "$0"
        financial_exposure = f"${shipment.Financial_Exposure:,.0f}" if shipment.Financial_Exposure else "$0"
        
        # Format boolean fields
        intent_to_claim = "✓" if shipment.Intent_To_Claim == 'YES' else "✗"
        formal_claim = "✓" if shipment.Formal_Claim_Received == 'YES' else "✗"
        
        # Format status
        settlement_status = ''
        if shipment.Settlement_Status == 'SETTLED':
            settlement_status = '✓'
        elif shipment.Settlement_Status == 'NOT_SETTLED':
            settlement_status = '✗'
        elif shipment.Settlement_Status == 'PARTIAL':
            settlement_status = '~'
        else:
            settlement_status = '-'
        
        status_symbol = ''
        if shipment.Status == 'OPEN':
            status_symbol = '●'
        elif shipment.Status == 'CLOSED':
            status_symbol = '✓'
        elif shipment.Status == 'PENDING':
            status_symbol = '⏳'
        elif shipment.Status == 'REJECTED':
            status_symbol = '✗'
        elif shipment.Status == 'UNDER_REVIEW':
            status_symbol = '◐'
        else:
            status_symbol = shipment.Status[:3] if shipment.Status else 'OPN'
        
        # Truncate shipment number for PDF to fit better
        shipment_no_display = shipment.Claim_No
        if len(shipment.Claim_No) > 15:
            shipment_no_display = shipment.Claim_No[:12] + '...'
        
        row = [
            shipment_no_display,  # Truncated for PDF display
            (shipment.Brand or '-')[:8] + '...' if shipment.Brand and len(shipment.Brand) > 8 else (shipment.Brand or '-'),
            (shipment.Claimant or '-')[:10] + '...' if shipment.Claimant and len(shipment.Claimant) > 10 else (shipment.Claimant or '-'),
            client_id,
            client_name[:12] + '...' if len(client_name) > 12 else client_name,
            intent_to_claim,
            intend_date,
            formal_claim,
            formal_date,
            claimed_amount,
            iscm_paid,
            carrier_paid,
            insurance_paid,
            shipment.Branch,
            total_savings,
            settlement_status,
            financial_exposure,
            status_symbol,
            closed_date
        ]
        data.append(row)
    
    # Create table with smaller font to fit all columns
    table = Table(data, repeatRows=1)
    
    # Add style to table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 8),  # Smaller header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),  # Header bottom padding
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Data alignment
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Data font
        ('FONTSIZE', (0, 1), (-1, -1), 6),  # Smaller data font size
        ('TOPPADDING', (0, 1), (-1, -1), 2),  # Data top padding
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),  # Data bottom padding
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Grid style
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),  # Box style
        # Alternate row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        # Right align amount columns
        ('ALIGN', (9, 1), (12, -1), 'RIGHT'),  # Value, ISCM, Carrier, Insurance
        ('ALIGN', (14, 1), (14, -1), 'RIGHT'),  # Savings
        ('ALIGN', (16, 1), (16, -1), 'RIGHT'),  # Exposure
    ]))
    
    # Add table to elements
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    pdf_data = buffer.getvalue()
    buffer.close()
    

    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
    response.write(pdf_data)
    
    return response


# =============================================================================
# IMPORT VIEWS - UPDATED FOR NEW CLIENT-SPECIFIC SHIPMENT IDs
# =============================================================================

@login_required(login_url='login')
def import_shipments(request):
    """Import shipment data from an Excel file, with detailed feedback on the process."""
    clear_messages(request)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file format. Please upload an Excel file.')
            return render(request, 'main/import_shipments.html')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            worksheet = wb.active
            skipped_entries, created_entries, error_entries = process_excel_data(worksheet)
            if created_entries == 0 and not skipped_entries and not error_entries:
                messages.info(request, 'No new entries were created. Check if the data is already up to date.')
            else:
                messages.success(request, f'Successfully created {created_entries} entries. Skipped {len(skipped_entries)} duplicate entries.')
                if error_entries:
                    messages.error(request, f'Errors occurred in {len(error_entries)} entries. {", ".join(error_entries)}')
        except Exception as e:
            messages.error(request, f'An error occurred while processing the file: {str(e)}')

        return render(request, 'main/import_shipments.html')
    return render(request, 'main/import_shipments.html')


def process_excel_data(worksheet):
    """Process Excel data and save valid entries to the database."""
    skipped_entries = []
    created_entries = 0
    error_entries = []
    
    # Updated column order for import:
    # 0: Shipment Number (can be blank - will auto-generate), 1: Brand, 2: Claimant, 
    # 3: Intent To Claim, 4: Intent To Claim Date, 5: Formal Claim, 6: Formal Claim Date, 
    # 7: Value, 8: Paid By ISCM, 9: Paid By Carrier, 10: Paid By Insurance, 11: Branch, 
    # 12: Total Savings, 13: Settled or Not Settled, 14: Financial Exposure, 15: Status
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:  # Skip completely empty rows
            continue
            
        # Try to get claim number from column 0, or leave blank for auto-generation
        claim_no = str(row[0]).strip() if row[0] else ""
        
        # Skip if claim number exists and is already in database
        if claim_no and Shipment.objects.filter(Claim_No=claim_no).exists():
            skipped_entries.append(claim_no)
            continue
            
        try:
            # Get Brand (column 1)
            brand = ""
            if len(row) > 1 and row[1]:
                brand = str(row[1]).strip()
            
            # Get Claimant (column 2) - REQUIRED for client identification
            claimant = ""
            if len(row) > 2 and row[2]:
                claimant = str(row[2]).strip()
            
            # Skip rows without claimant as we need it to identify/create client
            if not claimant:
                error_entries.append(f'Row {worksheet.iter_rows().__next__()}: Missing claimant name')
                continue
            
            # Get or create client based on claimant name
            client, created = Client.objects.get_or_create(
                name__iexact=claimant,
                defaults={'name': claimant}
            )
            
            # Handle date conversions for Intent To Claim Date (column 4)
            intend_date = None
            if len(row) > 4 and row[4]:
                if isinstance(row[4], datetime.date):
                    intend_date = row[4]
                elif isinstance(row[4], str):
                    try:
                        intend_date = datetime.datetime.strptime(row[4], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            intend_date = datetime.datetime.strptime(row[4], "%d/%m/%Y").date()
                        except ValueError:
                            intend_date = None
            
            # Handle date conversions for Formal Claim Date (column 6)
            formal_date = None
            if len(row) > 6 and row[6]:
                if isinstance(row[6], datetime.date):
                    formal_date = row[6]
                elif isinstance(row[6], str):
                    try:
                        formal_date = datetime.datetime.strptime(row[6], "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            formal_date = datetime.datetime.strptime(row[6], "%d/%m/%Y").date()
                        except ValueError:
                            formal_date = None
            
            # Handle numeric conversions safely
            # Value (column 7)
            claimed_amount = 0
            if len(row) > 7 and row[7]:
                if isinstance(row[7], (int, float)):
                    claimed_amount = float(row[7])
                elif isinstance(row[7], str) and row[7].strip():
                    clean_str = ''.join(c for c in row[7] if c.isdigit() or c == '.')
                    if clean_str:
                        claimed_amount = float(clean_str)
            
            # Paid By ISCM (column 8)
            iscm_amount = 0
            if len(row) > 8 and row[8]:
                if isinstance(row[8], (int, float)):
                    iscm_amount = float(row[8])
                elif isinstance(row[8], str) and row[8].strip():
                    clean_str = ''.join(c for c in row[8] if c.isdigit() or c == '.')
                    if clean_str:
                        iscm_amount = float(clean_str)
            
            # Paid By Carrier (column 9)
            carrier_amount = 0
            if len(row) > 9 and row[9]:
                if isinstance(row[9], (int, float)):
                    carrier_amount = float(row[9])
                elif isinstance(row[9], str) and row[9].strip():
                    clean_str = ''.join(c for c in row[9] if c.isdigit() or c == '.')
                    if clean_str:
                        carrier_amount = float(clean_str)
            
            # Paid By Insurance (column 10)
            insurance_amount = 0
            if len(row) > 10 and row[10]:
                if isinstance(row[10], (int, float)):
                    insurance_amount = float(row[10])
                elif isinstance(row[10], str) and row[10].strip():
                    clean_str = ''.join(c for c in row[10] if c.isdigit() or c == '.')
                    if clean_str:
                        insurance_amount = float(clean_str)
            
            # Total Savings (column 12)
            total_savings = 0
            if len(row) > 12 and row[12]:
                if isinstance(row[12], (int, float)):
                    total_savings = float(row[12])
                elif isinstance(row[12], str) and row[12].strip():
                    clean_str = ''.join(c for c in row[12] if c.isdigit() or c == '.')
                    if clean_str:
                        total_savings = float(clean_str)
            
            # Financial Exposure (column 14)
            financial_exposure = 0
            if len(row) > 14 and row[14]:
                if isinstance(row[14], (int, float)):
                    financial_exposure = float(row[14])
                elif isinstance(row[14], str) and row[14].strip():
                    clean_str = ''.join(c for c in row[14] if c.isdigit() or c == '.')
                    if clean_str:
                        financial_exposure = float(clean_str)
            
            # Convert Intent To Claim to YES/NO format (column 3)
            intent_claim = "NO"
            if len(row) > 3 and row[3]:
                if isinstance(row[3], bool):
                    intent_claim = "YES" if row[3] else "NO"
                elif isinstance(row[3], str):
                    if row[3].upper() in ["YES", "Y", "TRUE", "1"]:
                        intent_claim = "YES"
            
            # Convert Formal Claim to YES/NO format (column 5)
            formal_claim = "NO"
            if len(row) > 5 and row[5]:
                if isinstance(row[5], bool):
                    formal_claim = "YES" if row[5] else "NO"
                elif isinstance(row[5], str):
                    if row[5].upper() in ["YES", "Y", "TRUE", "1"]:
                        formal_claim = "YES"
            
            # Get branch value (column 11)
            branch = ""
            if len(row) > 11 and row[11]:
                branch = str(row[11]).strip()
                # Validate branch code
                if branch and branch not in [choice[0] for choice in Shipment.BRANCH_CHOICES]:
                    branch = ""  # Set to empty if invalid
            
            # Handle Settlement Status (column 13)
            settlement_status = None
            if len(row) > 13 and row[13]:
                settlement_value = str(row[13]).upper().strip()
                if settlement_value in ["SETTLED", "YES", "Y", "TRUE", "1"]:
                    settlement_status = "SETTLED"
                elif settlement_value in ["NOT SETTLED", "NOT_SETTLED", "NO", "N", "FALSE", "0"]:
                    settlement_status = "NOT_SETTLED"
                elif settlement_value in ["PARTIAL", "PARTIALLY SETTLED", "PARTIAL_SETTLED"]:
                    settlement_status = "PARTIAL"
            
            # Handle Status (column 15)
            status = "OPEN"  # Default status
            if len(row) > 15 and row[15]:
                status_value = str(row[15]).upper().strip()
                valid_statuses = [choice[0] for choice in Shipment.STATUS_CHOICES]
                if status_value in valid_statuses:
                    status = status_value
                elif status_value in ["OPEN", "PENDING", "CLOSED", "REJECTED", "UNDER_REVIEW"]:
                    status = status_value
            
            # Create shipment object - Claim_No will be auto-generated if blank
            shipment = Shipment(
                Claim_No=claim_no,  # Leave blank for auto-generation
                client=client,
                Brand=brand,
                Claimant=claimant,
                Intent_To_Claim=intent_claim,
                Intend_Claim_Date=intend_date,
                Formal_Claim_Received=formal_claim,
                Formal_Claim_Date_Received=formal_date,
                Claimed_Amount=claimed_amount,
                Amount_Paid_By_Awa=iscm_amount,
                Amount_Paid_By_Carrier=carrier_amount,
                Amount_Paid_By_Insurance=insurance_amount,
                Branch=branch,
                Total_Savings=total_savings,
                Settlement_Status=settlement_status,
                Financial_Exposure=financial_exposure,
                Status=status
            )
            shipment.save()  # Auto-generates Claim_No if blank
            created_entries += 1
            
        except Exception as e:
            error_entries.append(f'Row with Claimant {claimant}: {str(e)}')

    return skipped_entries, created_entries, error_entries


def api_inspections(request):
    """JSON API endpoint for Next.js frontend — returns inspection groups with their data."""
    from django.db.models import Exists, OuterRef, Subquery, F, Q, Count
    from ..models import InspectionGroup, FoodSafetyAgencyInspection

    def json_response(data, status=200):
        response = JsonResponse(data, status=status, safe=False)
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        return response

    if request.method == 'OPTIONS':
        return json_response({})

    try:
        show_duplicates = request.GET.get('show_duplicates') == 'true'
        show_undeliverable = request.GET.get('show_undeliverable') == 'true'
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        client_search = request.GET.get('client_search', '').strip()

        # Compute duplicate groups: same client_name + date + inspector + SAME products
        # Single query to get all candidate duplicate group IDs with their product signatures
        from collections import defaultdict
        _dup_triples = list(
            InspectionGroup.objects.values('client_name', 'date_of_inspection', 'inspector_name')
            .annotate(_gc=Count('id'))
            .filter(_gc__gt=1)
            .values_list('client_name', 'date_of_inspection', 'inspector_name')
        )
        _dup_group_ids = set()
        if _dup_triples:
            # Build Q filter for all candidate combos at once
            _combo_q = Q()
            for _dc, _dd, _di in _dup_triples:
                _combo_q |= Q(client_name=_dc, date_of_inspection=_dd, inspector_name=_di)
            # Single query: fetch all candidate groups + their inspections
            _candidate_groups = list(
                InspectionGroup.objects.filter(_combo_q)
                .prefetch_related('inspections')
            )
            # Group by (client_name, date, inspector) then by product signature
            _combo_map = defaultdict(list)  # (client, date, inspector) -> [(group_id, signature)]
            for _cg in _candidate_groups:
                _products = tuple(sorted(
                    (_i.commodity or '', _i.product_name or '')
                    for _i in _cg.inspections.all()
                ))
                _combo_map[(_cg.client_name, _cg.date_of_inspection, _cg.inspector_name)].append((_cg.id, _products))
            # Find true duplicates: groups with identical product signatures
            for _combo_key, _entries in _combo_map.items():
                _sig_map = defaultdict(list)
                for _gid, _sig in _entries:
                    _sig_map[_sig].append(_gid)
                for _sig, _gids in _sig_map.items():
                    if len(_gids) > 1:
                        _dup_group_ids.update(_gids)

        _dup_q = Q(id__in=_dup_group_ids) if _dup_group_ids else Q(pk__in=[])

        insp = FoodSafetyAgencyInspection.objects.filter(inspection_group_id=OuterRef('pk'))
        groups_qs = (
            InspectionGroup.objects
            .select_related('client')
            .prefetch_related('inspections')
            .annotate(
                is_occurrence_report=Exists(insp.filter(is_occurrence_report=True)),
                first_approved=Subquery(insp.order_by('id').values('approved_status')[:1]),
                first_sent=Subquery(insp.order_by('id').values('sent_date')[:1]),
                first_sent_by_id=Subquery(insp.filter(sent_by__isnull=False).order_by('id').values('sent_by')[:1]),
            )
            .order_by('-date_of_inspection')
        )

        # Inspector role: only show their own inspections
        # Inspector Manager: sees ALL inspections (like admin)
        _is_auth = hasattr(request, 'user') and request.user.is_authenticated
        _user_role = getattr(request.user, 'role', '') if _is_auth else 'anonymous'
        print(f"[INSPECTIONS_API] authenticated={_is_auth}, user={request.user if _is_auth else 'anon'}, role={_user_role}")
        if _is_auth:
            user_role = _user_role
            if user_role == 'inspector':
                from ..models import InspectorMapping
                inspector_name = request.user.get_full_name() or request.user.username
                inspector_q = Q(inspector_name__iexact=inspector_name)
                mapping = InspectorMapping.objects.filter(inspector_name__iexact=inspector_name).first()
                if not mapping:
                    mapping = InspectorMapping.objects.filter(inspector_name__iexact=request.user.username).first()
                if mapping and mapping.inspector_id:
                    inspector_group_ids = FoodSafetyAgencyInspection.objects.filter(
                        inspector_id=mapping.inspector_id
                    ).values_list('inspection_group_id', flat=True).distinct()
                    inspector_q = inspector_q | Q(id__in=list(inspector_group_ids))
                groups_qs = groups_qs.filter(inspector_q)
            # inspector_manager sees all inspections - no filtering needed

        # Server-side inspector filter
        filter_inspector = request.GET.get('inspector', '').strip()
        if filter_inspector:
            groups_qs = groups_qs.filter(inspector_name__iexact=filter_inspector)

        # Server-side corporate group filter
        filter_corp_group = request.GET.get('corporate_group', '').strip()
        if filter_corp_group:
            groups_qs = groups_qs.filter(corporate_group__iexact=filter_corp_group)

        if date_from:
            groups_qs = groups_qs.filter(date_of_inspection__gte=date_from)
        if date_to:
            groups_qs = groups_qs.filter(date_of_inspection__lte=date_to)
        if client_search:
            groups_qs = groups_qs.filter(client_name__icontains=client_search)

        # Count duplicate groups for the badge
        duplicate_groups_count = groups_qs.filter(_dup_q).count() if _dup_q else 0

        # Undeliverable email check — scans ALL inspections globally (ignores date filters).
        # Only does the Graph API call when show_undeliverable=true; otherwise uses cached count.
        undeliverable_count = 0
        _undel_q = None
        _bounced_emails_list = []
        if show_undeliverable:
            # Full check: fetch bounced emails and resolve to client names
            try:
                from ..graph_inbox_reader import fetch_undeliverable_emails
                from ..models import Client as _UndelClient
                from django.core.cache import cache as _dj_cache
                _bounced_emails = fetch_undeliverable_emails()
                if _bounced_emails:
                    _bounced_lower = {e.lower() for e in _bounced_emails}
                    _bounced_emails_list = sorted(_bounced_lower)
                    _undel_clients = _UndelClient.objects.all().only('name', 'email', 'manual_email').prefetch_related('additional_emails')
                    _undeliverable_client_names = set()
                    for _uc in _undel_clients:
                        _uc_emails = set()
                        if _uc.email:
                            for _e in _uc.email.replace(';', ',').split(','):
                                _e = _e.strip().lower()
                                if _e:
                                    _uc_emails.add(_e)
                        if _uc.manual_email:
                            for _e in _uc.manual_email.replace(';', ',').split(','):
                                _e = _e.strip().lower()
                                if _e:
                                    _uc_emails.add(_e)
                        for _ae in _uc.additional_emails.all():
                            if _ae.email:
                                _uc_emails.add(_ae.email.strip().lower())
                        if _uc_emails & _bounced_lower:
                            _undeliverable_client_names.add(_uc.name)
                    if _undeliverable_client_names:
                        _undel_q = Q()
                        for _ucn in _undeliverable_client_names:
                            _undel_q |= Q(client_name=_ucn)
                        # Count against ALL inspections (no date filter) so badge is global
                        _all_groups = InspectionGroup.objects.all()
                        undeliverable_count = _all_groups.filter(_undel_q).count()
                        # Cache the count and the bounced emails list
                        _dj_cache.set('inspections_undeliverable_count', undeliverable_count, 3600)
                        _dj_cache.set('inspections_bounced_emails', _bounced_emails_list, 3600)
            except Exception as _e:
                print(f"[WARN] api_inspections undeliverable check failed: {_e}")
        else:
            # Badge count only — use cached value (avoid Graph API call)
            try:
                from django.core.cache import cache as _dj_cache
                _cached_count = _dj_cache.get('inspections_undeliverable_count')
                if _cached_count is not None:
                    undeliverable_count = _cached_count
            except Exception:
                pass

        # Filter to only undeliverable if requested — apply against UNFILTERED queryset
        # so we see ALL undeliverable inspections regardless of date range
        if show_undeliverable and _undel_q is not None and undeliverable_count > 0:
            # Reset to all inspections (remove date filters) then apply undeliverable filter
            groups_qs = (
                InspectionGroup.objects
                .select_related('client')
                .prefetch_related('inspections')
                .annotate(
                    is_occurrence_report=Exists(insp.filter(is_occurrence_report=True)),
                    first_approved=Subquery(insp.order_by('id').values('approved_status')[:1]),
                    first_sent=Subquery(insp.filter(sent_date__isnull=False).order_by('-id').values('sent_date')[:1]),
                    first_sent_by_id=Subquery(insp.filter(sent_by__isnull=False).order_by('-id').values('sent_by')[:1]),
                )
                .filter(_undel_q)
                .order_by('-date_of_inspection')
            )

        # Filter to only duplicates if requested
        if show_duplicates:
            if _dup_group_ids:
                groups_qs = groups_qs.filter(_dup_q)
                # Sort by client_name then date so duplicates appear next to each other
                groups_qs = groups_qs.order_by('client_name', '-date_of_inspection', 'id')
            else:
                groups_qs = groups_qs.none()

        # Server-side pagination
        _total_count = groups_qs.count()
        _page = int(request.GET.get('page', '1'))
        _page_size = min(int(request.GET.get('page_size', '50')), 200)
        _offset = (_page - 1) * _page_size
        groups = groups_qs[_offset:_offset + _page_size]

        # ── Batch filesystem scan (only for current page's inspection IDs) ──
        _file_map = {}  # (insp_id_str, category) -> bool
        _FILE_CATEGORIES = {'lab', 'coa', 'composition', 'compliance',
                            'occurrence', 'retest', 'other', 'lab_form',
                            'rfi', 'invoice'}
        # Collect all inspection IDs for the current page only
        _page_insp_ids = set()
        for g in groups:
            for p in g.inspections.all():
                _page_insp_ids.add(str(p.id))
        _docs_root = os.path.join(settings.MEDIA_ROOT, 'docs')
        if os.path.isdir(_docs_root) and _page_insp_ids:
            try:
                for _cid in os.listdir(_docs_root):
                    _client_dir = os.path.join(_docs_root, _cid)
                    if not os.path.isdir(_client_dir):
                        continue
                    for _iid in os.listdir(_client_dir):
                        if _iid not in _page_insp_ids:
                            continue
                        _insp_dir = os.path.join(_client_dir, _iid)
                        if not os.path.isdir(_insp_dir):
                            continue
                        for _cat in os.listdir(_insp_dir):
                            if _cat in _FILE_CATEGORIES:
                                _cat_path = os.path.join(_insp_dir, _cat)
                                if os.path.isdir(_cat_path) and os.listdir(_cat_path):
                                    _file_map[(_iid, _cat)] = True
            except OSError:
                pass

        def _has_file(insp_id, category):
            return _file_map.get((str(insp_id), category), False)

        # Build user-id → name lookup for sent_by
        _sent_by_ids = set(g.first_sent_by_id for g in groups if getattr(g, 'first_sent_by_id', None))
        _user_name_map = {}
        if _sent_by_ids:
            from django.contrib.auth import get_user_model
            _U = get_user_model()
            for _u in _U.objects.filter(id__in=_sent_by_ids).only('id', 'first_name', 'last_name', 'username'):
                _user_name_map[_u.id] = f"{_u.first_name} {_u.last_name}".strip() or _u.username

        results = []
        for g in groups:
            inspections = list(g.inspections.all())
            first_insp = inspections[0] if inspections else None
            insp_ids = [p.id for p in inspections]

            # Group-level file flags (no per-product detail needed for list)
            has_rfi = any(_has_file(pid, 'rfi') for pid in insp_ids)
            has_invoice = any(_has_file(pid, 'invoice') for pid in insp_ids)
            has_lab = any(_has_file(pid, 'lab') or _has_file(pid, 'coa') for pid in insp_ids)
            has_compliance = any(_has_file(pid, 'compliance') for pid in insp_ids)
            has_lab_form = any(_has_file(pid, 'lab_form') for pid in insp_ids)
            # Compliance: matches frontend badge logic
            # Product is "assessed" if it has compliance or composition file uploaded
            # NON_COMPLIANT if any product is non-compliant (with file proof)
            # COMPLIANT if at least one product is assessed+compliant and none are non-compliant
            # PENDING if no products are assessed
            _any_non_compliant = False
            _any_compliant = False
            for p in inspections:
                has_compliance_doc = _has_file(p.id, 'compliance')
                has_composition_doc = _has_file(p.id, 'composition')
                is_assessed = has_compliance_doc or has_composition_doc
                if is_assessed:
                    if getattr(p, 'is_product_compliant', True):
                        _any_compliant = True
                    else:
                        _any_non_compliant = True
                # No file = N/A, ignore is_product_compliant value

            if _any_non_compliant:
                _compliance_status = 'NON_COMPLIANT'
            elif _any_compliant:
                _compliance_status = 'COMPLIANT'
            else:
                _compliance_status = 'PENDING'

            # Build products — compact for list view
            products = []
            for p in inspections:
                prod = {
                    'id': p.id,
                    'commodity': p.commodity or '',
                    'product_name': p.product_name or '',
                    'product_class': p.product_class or '',
                    'is_product_compliant': bool(p.is_product_compliant),
                    'is_sample_taken': bool(p.is_sample_taken) if p.is_sample_taken is not None else False,
                    'needs_retest': p.needs_retest or '',
                    'lab': p.get_lab_display() if p.lab else '',
                }
                # Only include test flags and file flags if they are True (saves bytes)
                if p.dna: prod['dna'] = True
                if p.fat: prod['fat'] = True
                if p.protein: prod['protein'] = True
                if p.calcium: prod['calcium'] = True
                if p.is_direction_present_for_this_inspection: prod['is_direction_present_for_this_inspection'] = True
                if _has_file(p.id, 'lab') or _has_file(p.id, 'coa'): prod['coa_uploaded'] = True
                if _has_file(p.id, 'composition'): prod['composition_uploaded'] = True
                if _has_file(p.id, 'compliance'): prod['compliance_uploaded'] = True
                if _has_file(p.id, 'occurrence'): prod['occurrence_uploaded'] = True
                if _has_file(p.id, 'retest'): prod['retest_uploaded'] = True
                if _has_file(p.id, 'other'): prod['other_uploaded'] = True
                if _has_file(p.id, 'lab_form'): prod['lab_form_uploaded'] = True
                products.append(prod)

            # Build fallback_group_id string
            client_slug = re.sub(r'[^a-zA-Z0-9]', '_', g.client_name or 'Unknown')
            client_slug = re.sub(r'_+', '_', client_slug).strip('_')
            date_str = g.date_of_inspection.strftime('%Y%m%d') if g.date_of_inspection else '00000000'
            fallback_group_id = f"{client_slug}_{date_str}_g{g.id}"

            results.append({
                'id': g.id,
                'group_id': fallback_group_id,
                'client_name': g.client_name or '',
                'town': g.town or '',
                'inspector_name': g.inspector_name or '',
                'date_of_inspection': g.date_of_inspection.isoformat() if g.date_of_inspection else None,
                'approved_status': g.first_approved or 'PENDING',
                'sent_date': g.first_sent.isoformat() if g.first_sent else None,
                'sent_by_name': _user_name_map.get(g.first_sent_by_id, '') if getattr(g, 'first_sent_by_id', None) else '',
                'is_occurrence_report': g.is_occurrence_report,
                'has_rfi': has_rfi,
                'has_invoice': has_invoice,
                'has_lab': has_lab,
                'has_compliance': has_compliance,
                'email': '; '.join(filter(None, [
                    g.client.email if g.client else '',
                    g.additional_email or '',
                ])).strip('; ') or '',
                'km_traveled': float(g.km_traveled) if g.km_traveled else 0,
                'hours': float(g.hours) if g.hours else 0,
                'group_type': g.group_type or '',
                'facility_type': g.facility_type or '',
                'corporate_group': g.corporate_group or '',
                'internal_account_code': first_insp.internal_account_code if first_insp and first_insp.internal_account_code else '',
                'registration_code': first_insp.registration_code if first_insp and first_insp.registration_code else '',
                'physical_address': first_insp.physical_address if first_insp and first_insp.physical_address else '',
                'telephone': first_insp.telephone if first_insp and first_insp.telephone else '',
                'time_of_visit': first_insp.time_of_visit.strftime('%H:%M') if first_insp and first_insp.time_of_visit else '',
                'comment': (first_insp.occurrence_description if first_insp and first_insp.occurrence_description else '') or g.comment or (first_insp.comment if first_insp else '') or '',
                'has_lab_form': has_lab_form,
                'inspection_compliance_status': _compliance_status,
                'products': products,
            })

        import math as _math

        # Get ALL unique inspectors for filter dropdown - only actual inspectors
        from django.contrib.auth import get_user_model as _gum
        _User = _gum()
        _non_inspector_names = set()
        for _u in _User.objects.exclude(role__in=['inspector', 'inspector_manager']):
            _full = f"{_u.first_name} {_u.last_name}".strip()
            if _full: _non_inspector_names.add(_full)
            _non_inspector_names.add(_u.username)
        _non_inspector_names.update(['API User', 'Test Inspector', 'admin'])
        _all_inspectors = sorted(set(
            n for n in InspectionGroup.objects.exclude(inspector_name__isnull=True).exclude(inspector_name='')
            .values_list('inspector_name', flat=True)
            if n not in _non_inspector_names
        ))
        _all_corp_groups = sorted(set(
            InspectionGroup.objects.exclude(corporate_group__isnull=True).exclude(corporate_group='')
            .values_list('corporate_group', flat=True)
        ))
        _all_group_types = sorted(set(
            InspectionGroup.objects.exclude(group_type__isnull=True).exclude(group_type='')
            .values_list('group_type', flat=True)
        ))

        _resp = {
            'count': _total_count,
            'results': results,
            'page': _page,
            'page_size': _page_size,
            'total_pages': _math.ceil(_total_count / _page_size) if _page_size else 1,
            'duplicate_groups_count': duplicate_groups_count,
            'show_duplicates': show_duplicates,
            'undeliverable_count': undeliverable_count,
            'all_inspectors': _all_inspectors,
            'all_corporate_groups': _all_corp_groups,
            'all_group_types': _all_group_types,
        }
        # Include bounced emails list so frontend can highlight them
        if show_undeliverable and _bounced_emails_list:
            _resp['bounced_emails'] = _bounced_emails_list
        return json_response(_resp)

    except Exception as e:
        return json_response({'error': str(e), 'count': 0, 'results': []}, status=500)


def api_clients(request):
    """JSON API endpoint for Next.js frontend — returns clients with filtering, sorting, and pagination."""
    from django.db.models import Q
    from django.core.paginator import Paginator
    from ..models import Client

    def json_response(data, status=200):
        response = JsonResponse(data, status=status, safe=False)
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        return response

    if request.method == 'OPTIONS':
        return json_response({})

    try:
        # Get filter parameters
        client_id = request.GET.get('client_id', '').strip()
        client_name = request.GET.get('client_name', '').strip()
        corporate_group = [v.strip() for v in request.GET.get('corporate_group', '').split(',') if v.strip()]
        commodity = [v.strip() for v in request.GET.get('commodity', '').split(',') if v.strip()]
        facility_type = [v.strip() for v in request.GET.get('facility_type', '').split(',') if v.strip()]
        facility_code = [v.strip() for v in request.GET.get('facility_code', '').split(',') if v.strip()]
        province = [v.strip() for v in request.GET.get('province', '').split(',') if v.strip()]
        account_code = request.GET.get('account_code', '').strip()
        group_type = [v.strip() for v in request.GET.get('group_type', '').split(',') if v.strip()]

        # Get sorting and pagination parameters
        sort_by = request.GET.get('sort_by', 'name')
        sort_order = request.GET.get('sort_order', 'asc')
        per_page = int(request.GET.get('per_page', '100'))
        page_number = int(request.GET.get('page', '1'))

        # Validate sorting parameters
        valid_sort_fields = ['name', 'client_id', 'facility_type', 'corporate_group', 'town', 'email', 'group_type']
        if sort_by not in valid_sort_fields:
            sort_by = 'name'

        # Build order_by string
        order_by_field = f'-{sort_by}' if sort_order == 'desc' else sort_by

        # Start with base query
        clients_query = Client.objects.all()

        # Apply filters
        if client_id:
            clients_query = clients_query.filter(client_id__icontains=client_id)
        if client_name:
            clients_query = clients_query.filter(name__icontains=client_name)
        if corporate_group:
            clients_query = clients_query.filter(corporate_group__in=corporate_group)
        if facility_type:
            clients_query = clients_query.filter(facility_type__in=facility_type)
        if group_type:
            clients_query = clients_query.filter(group_type__in=group_type)
        if province:
            clients_query = clients_query.filter(town__in=province)
        if account_code:
            clients_query = clients_query.filter(internal_account_code__icontains=account_code)
        if facility_code:
            q = Q()
            for fc in facility_code:
                q |= Q(internal_account_code__istartswith=fc)
            clients_query = clients_query.filter(q)
        if commodity:
            q = Q()
            for c in commodity:
                q |= Q(internal_account_code__icontains=f'-{c}-')
            clients_query = clients_query.filter(q)

        # Apply sorting
        clients_query = clients_query.order_by(order_by_field)

        # Get total count
        total_count = clients_query.count()

        # Paginate
        paginator = Paginator(clients_query, per_page)
        page_obj = paginator.get_page(page_number)

        # Build client results
        clients_list = []
        for client in page_obj.object_list:
            clients_list.append({
                'client_id': client.client_id,
                'name': client.name or '',
                'facility_type': client.facility_type or '',
                'town': client.town or '',
                'corporate_group': client.corporate_group or '',
                'internal_account_code': client.internal_account_code or '',
                'email': client.email or '',
                'phone_number': '',
                'group_type': client.group_type or '',
            })

        # Get unique values for dropdown filters
        corporate_groups = list(
            Client.objects.exclude(corporate_group__isnull=True).exclude(corporate_group='')
            .values_list('corporate_group', flat=True).distinct().order_by('corporate_group')
        )
        facility_types = list(
            Client.objects.exclude(facility_type__isnull=True).exclude(facility_type='')
            .values_list('facility_type', flat=True).distinct().order_by('facility_type')
        )
        group_types = list(
            Client.objects.exclude(group_type__isnull=True).exclude(group_type='')
            .values_list('group_type', flat=True).distinct().order_by('group_type')
        )
        provinces = list(
            Client.objects.exclude(town__isnull=True).exclude(town='')
            .values_list('town', flat=True).distinct().order_by('town')
        )

        return json_response({
            'clients': clients_list,
            'total_count': total_count,
            'page': page_obj.number,
            'num_pages': paginator.num_pages,
            'per_page': per_page,
            'filters': {
                'corporate_groups': corporate_groups,
                'facility_types': facility_types,
                'group_types': group_types,
                'provinces': provinces,
            }
        })

    except Exception as e:
        return json_response({'error': str(e), 'clients': [], 'total_count': 0}, status=500)


@login_required
def get_inspection_fees(request):
    """Get all inspection fees with current and historical rate information"""
    from ..models import InspectionFee

    fees = InspectionFee.objects.all()
    fees_data = []

    for fee in fees:
        # Get the most recent history entry to find current effective date
        latest_history = fee.history.order_by('-effective_date').first()

        fee_dict = {
            'id': fee.id,
            'fee_code': fee.fee_code,
            'fee_name': fee.fee_name,
            'rate': float(fee.rate),
            'description': fee.description,
            'last_updated': fee.last_updated.isoformat() if fee.last_updated else None,
            'effective_date': latest_history.effective_date.isoformat() if latest_history else None,
            'has_history': fee.history.exists(),
            'history_count': fee.history.count()
        }
        fees_data.append(fee_dict)

    return JsonResponse({'fees': fees_data})


@login_required
@require_POST
def update_inspection_fees(request):
    """
    Update inspection fees with versioning support.
    Creates FeeHistory entries instead of directly overwriting rates.
    """
    import json
    from ..models import InspectionFee, FeeHistory
    from datetime import date
    from decimal import Decimal

    try:
        data = json.loads(request.body)
        fees_data = data.get('fees', [])
        effective_date_str = data.get('effective_date')  # Optional: allow setting effective date

        # Parse effective date or default to today
        if effective_date_str:
            try:
                effective_date = date.fromisoformat(effective_date_str)
            except (ValueError, TypeError):
                effective_date = date.today()
        else:
            effective_date = date.today()

        updated_count = 0
        errors = []

        for fee_data in fees_data:
            fee_id = fee_data.get('id')
            new_rate = fee_data.get('rate')

            if fee_id and new_rate is not None:
                try:
                    fee = InspectionFee.objects.get(id=fee_id)
                    new_rate_decimal = Decimal(str(new_rate))

                    # Check if rate actually changed
                    if fee.rate != new_rate_decimal:
                        # Create a new FeeHistory entry
                        FeeHistory.objects.create(
                            fee=fee,
                            rate=new_rate_decimal,
                            effective_date=effective_date,
                            created_by=request.user,
                            notes=fee_data.get('notes', '')  # Optional notes about the change
                        )

                        # Update the current rate on the fee
                        fee.rate = new_rate_decimal
                        fee.updated_by = request.user
                        fee.save()

                        updated_count += 1

                except InspectionFee.DoesNotExist:
                    errors.append(f"Fee with ID {fee_id} not found")
                except Exception as e:
                    errors.append(f"Error updating fee {fee_id}: {str(e)}")

        response_data = {
            'success': True,
            'message': f'Updated {updated_count} fees successfully',
            'updated_count': updated_count,
            'effective_date': effective_date.isoformat()
        }

        if errors:
            response_data['warnings'] = errors

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)


@login_required
def get_inspection_fee_history(request):
    """Get complete history of all fee changes"""
    from ..models import FeeHistory

    # Get all fee history ordered by effective date (most recent first)
    history = FeeHistory.objects.select_related('fee', 'created_by').order_by('-effective_date', '-created_at')

    history_data = []
    for record in history:
        # Get the previous rate by finding the next older history record for the same fee
        previous_history = FeeHistory.objects.filter(
            fee=record.fee,
            effective_date__lt=record.effective_date
        ).order_by('-effective_date').first()

        history_dict = {
            'id': record.id,
            'fee_name': record.fee.fee_name,
            'fee_code': record.fee.fee_code,
            'rate': float(record.rate),
            'previous_rate': float(previous_history.rate) if previous_history else None,
            'effective_date': record.effective_date.isoformat(),
            'created_at': record.created_at.isoformat(),
            'created_by': record.created_by.username if record.created_by else None,
            'notes': record.notes or ''
        }
        history_data.append(history_dict)

    return JsonResponse({'history': history_data})


# =============================================================================
# CLIENT ALLOCATION API ENDPOINTS (No login required - for Next.js frontend)
# =============================================================================

@_csrf_exempt
def api_client_add(request):
    """API endpoint to add a new client allocation record. No login required."""
    from ..models import ClientAllocation
    from django.db.models import Max
    import json

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'POST required'}, status=405))

    try:
        data = json.loads(request.body)

        # Auto-generate next Client ID
        max_id = ClientAllocation.objects.aggregate(Max('client_id'))['client_id__max']
        client_id = (max_id or 0) + 1

        business_name = data.get('business_name', '').strip()
        facility_type = data.get('facility_type', '').strip()
        group_type = data.get('group_type', '').strip()
        commodity = data.get('commodity', '').strip()
        province = data.get('province', '').strip()
        corporate_group = data.get('corporate_group', '').strip()
        allocated = data.get('allocated', 'no') == 'yes'
        representative_email = data.get('representative_email', '').strip()
        phone_number = data.get('phone_number', '').strip()

        if not business_name:
            return _cors(JsonResponse({'success': False, 'error': 'Business name is required'}))

        # Auto-detect corporate group if not provided
        if not corporate_group:
            from .core_views import detect_corporate_group
            corporate_group = detect_corporate_group(business_name)

        # Generate internal account code
        corporate_group_codes = {
            'Not Applicable (None)': 'NA', 'Pick n Pay - Franchise': 'PNP-F',
            'Pick n Pay - Corporate': 'PNP-C', 'Fruit & Veg': 'FNV', 'OK Foods': 'OK',
            'Checkers': 'CHK', 'Spar': 'SPR', 'SuperSpar': 'SSPR', 'Spar - Northrand': 'SPR-N',
            'Shoprite': 'SHO', 'Massmart': 'MAS', 'Chester Butcheries': 'CHE', 'Boxer': 'BOX',
            'Food Lovers Market': 'FLM', 'Cambridge': 'CAM', 'Woolworths': 'WOO',
            'Jwayelani': 'JWA', 'Usave': 'USA', 'Other (Unlisted Group)': 'OTH',
            'OBC': 'OBC', 'Roots': 'ROO', 'Meat World': 'MEA', 'Quantum Foods Nulaid': 'QFN',
            'Bluff Meat Supply': 'BMS', 'Eat Sum Meat': 'ESM', 'Waltloo Meat and Chicken': 'WMC',
            'Choppies': 'CHO', 'Econo Foods': 'ECO', 'Makro': 'MAK', 'Boma Vleismark': 'BOM',
            'Eskort': 'ESK', 'Nesta Foods': 'NES',
        }
        part1 = facility_type[:2].upper() if facility_type else '-'
        part2 = 'IND'
        if group_type == 'Corporate Store':
            part2 = 'COR'
        elif group_type == 'Franchise Store':
            part2 = 'FRN'
        part3 = commodity if commodity in ['PMP', 'RAW', 'EGG', 'PLT'] else 'OTH'
        part4 = corporate_group_codes.get(corporate_group, '-')
        part5 = str(client_id).zfill(4)
        internal_account_code = f"{part1}-{part2}-{part3}-{part4}-{part5}"

        ClientAllocation.objects.create(
            client_id=client_id,
            eclick_name=business_name,
            facility_type=facility_type,
            group_type=group_type,
            commodity=commodity,
            province=province,
            corporate_group=corporate_group,
            internal_account_code=internal_account_code,
            allocated=allocated,
            representative_email=representative_email,
            phone_number=phone_number,
            manually_added=True,
        )

        # Also create a Client record so it shows in the clients list
        from ..models import Client
        Client.objects.get_or_create(
            name=business_name,
            defaults={
                'facility_type': facility_type,
                'group_type': group_type,
                'corporate_group': corporate_group,
                'town': province,
                'internal_account_code': internal_account_code,
                'email': representative_email,
            }
        )

        return _cors(JsonResponse({
            'success': True,
            'message': f'Client {client_id} - {business_name} added successfully!',
            'client_id': client_id,
        }))

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}))


@_csrf_exempt
def api_client_edit(request):
    """API endpoint to edit an existing client allocation record. No login required."""
    from ..models import ClientAllocation
    import json

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'POST required'}, status=405))

    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')

        if not client_id:
            return _cors(JsonResponse({'success': False, 'error': 'Missing client_id'}))

        allocation = ClientAllocation.objects.get(client_id=client_id)

        allocation.eclick_name = data.get('business_name', allocation.eclick_name)
        allocation.facility_type = data.get('facility_type', allocation.facility_type)
        allocation.group_type = data.get('group_type', allocation.group_type)
        allocation.commodity = data.get('commodity', allocation.commodity)
        allocation.province = data.get('province', allocation.province)

        corporate_group = data.get('corporate_group', '').strip()
        if not corporate_group:
            from .core_views import detect_corporate_group
            corporate_group = detect_corporate_group(allocation.eclick_name)
        allocation.corporate_group = corporate_group

        allocated_val = data.get('allocated', '')
        if allocated_val:
            allocation.allocated = allocated_val == 'yes'

        allocation.representative_email = data.get('representative_email', allocation.representative_email)
        allocation.phone_number = data.get('phone_number', allocation.phone_number)

        active_status = data.get('active_status', '')
        if active_status:
            allocation.active_status = active_status

        allocation.save()

        return _cors(JsonResponse({
            'success': True,
            'message': f'Client {client_id} - {allocation.eclick_name} updated successfully!',
        }))

    except ClientAllocation.DoesNotExist:
        return _cors(JsonResponse({'success': False, 'error': f'Client {client_id} not found'}))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}))


@_csrf_exempt
def api_client_delete(request):
    """API endpoint to delete a client allocation record. No login required."""
    from ..models import ClientAllocation
    import json

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'POST required'}, status=405))

    try:
        data = json.loads(request.body)
        client_id = data.get('client_id')

        if not client_id:
            return _cors(JsonResponse({'success': False, 'error': 'Missing client_id'}))

        deleted_count = ClientAllocation.objects.filter(client_id=client_id).delete()[0]

        if deleted_count > 0:
            return _cors(JsonResponse({
                'success': True,
                'message': f'Client {client_id} deleted successfully',
            }))
        else:
            return _cors(JsonResponse({'success': False, 'error': 'Client not found'}))

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}))


def api_dropdown_options(request):
    """API endpoint to get or add dropdown options. No login required."""
    from ..models import Client, ClientAllocation, ClientDropdownOption
    from django.db.models import Count, Q
    import json

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    if request.method == 'GET':
        def merge_options(db_qs, field, custom_qs):
            db_vals = {item[field]: item['count'] for item in db_qs}
            for opt in custom_qs:
                if opt.value not in db_vals:
                    db_vals[opt.value] = 0
            return sorted([{'value': k, 'count': v} for k, v in db_vals.items()], key=lambda x: x['value'])

        facility_types_qs = Client.objects.values('facility_type').annotate(count=Count('id')).filter(~Q(facility_type='') & ~Q(facility_type__isnull=True))
        corporate_groups_qs = Client.objects.values('corporate_group').annotate(count=Count('id')).filter(~Q(corporate_group='') & ~Q(corporate_group__isnull=True))
        group_types_qs = Client.objects.values('group_type').annotate(count=Count('id')).filter(~Q(group_type='') & ~Q(group_type__isnull=True))

        custom_ft = ClientDropdownOption.objects.filter(field_type='facility_type')
        custom_cg = ClientDropdownOption.objects.filter(field_type='corporate_group')
        custom_gt = ClientDropdownOption.objects.filter(field_type='group_type')

        return _cors(JsonResponse({
            'facility_types': merge_options(facility_types_qs, 'facility_type', custom_ft),
            'corporate_groups': merge_options(corporate_groups_qs, 'corporate_group', custom_cg),
            'group_types': merge_options(group_types_qs, 'group_type', custom_gt),
        }))

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            field_type = data.get('field_type', '').strip()
            value = data.get('value', '').strip()

            if not field_type or not value:
                return _cors(JsonResponse({'success': False, 'error': 'Missing field_type or value'}))

            valid_fields = ['facility_type', 'corporate_group', 'group_type']
            if field_type not in valid_fields:
                return _cors(JsonResponse({'success': False, 'error': 'Invalid field type'}))

            obj, created = ClientDropdownOption.objects.get_or_create(field_type=field_type, value=value)
            if not created:
                return _cors(JsonResponse({'success': False, 'error': f'"{value}" already exists'}))

            return _cors(JsonResponse({'success': True, 'message': f'Added "{value}" to {field_type}'}))

        except Exception as e:
            return _cors(JsonResponse({'success': False, 'error': str(e)}))

    return _cors(JsonResponse({'success': False, 'error': 'Invalid method'}, status=405))


def api_dropdown_option_delete(request):
    """API endpoint to delete a dropdown option. No login required."""
    from ..models import Client, ClientDropdownOption
    import json

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'POST required'}, status=405))

    try:
        data = json.loads(request.body)
        field_type = data.get('field_type')
        value = data.get('value')

        if not field_type or not value:
            return _cors(JsonResponse({'success': False, 'error': 'Missing field_type or value'}))

        field_map = {
            'facility_type': 'facility_type',
            'corporate_group': 'corporate_group',
            'group_type': 'group_type',
        }

        if field_type not in field_map:
            return _cors(JsonResponse({'success': False, 'error': 'Invalid field type'}))

        field_name = field_map[field_type]
        updated_count = Client.objects.filter(**{field_name: value}).update(**{field_name: ''})
        ClientDropdownOption.objects.filter(field_type=field_type, value=value).delete()

        return _cors(JsonResponse({
            'success': True,
            'updated_count': updated_count,
        }))

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}))


@_csrf_exempt
def api_users(request):
    """API endpoint for user management from Next.js frontend. No login required."""
    from django.contrib.auth import get_user_model
    from django.contrib.auth.hashers import make_password
    from ..models import InspectorMapping, InspectorSalary, InspectorManagerAllocation
    import json

    User = get_user_model()

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({}))

    # ── GET: return all users ──
    if request.method == 'GET':
        try:
            users = User.objects.all().order_by('username')
            # Pre-fetch all inspector mappings to avoid N+1 queries
            mapping_lookup = {m.inspector_name.lower(): m.inspector_id for m in InspectorMapping.objects.all()}
            users_list = []
            for u in users:
                role = getattr(u, 'role', 'inspector') or 'inspector'
                inspector_id = None
                if role == 'inspector':
                    key = f'{u.first_name} {u.last_name}'.lower()
                    inspector_id = mapping_lookup.get(key)

                users_list.append({
                    'id': u.id,
                    'username': u.username,
                    'email': u.email,
                    'first_name': u.first_name,
                    'last_name': u.last_name,
                    'role': role,
                    'is_active': u.is_active,
                    'inspector_id': inspector_id,
                })

            # Build inspector_users list (users with role='inspector')
            inspector_users = [
                {'id': u['id'], 'username': u['username'], 'first_name': u['first_name'], 'last_name': u['last_name']}
                for u in users_list if u['role'] == 'inspector'
            ]

            # Build manager_allocations map: manager_id -> list of inspector user ids
            manager_allocations = {}
            for alloc in InspectorManagerAllocation.objects.all():
                mid = alloc.manager_id
                if mid not in manager_allocations:
                    manager_allocations[mid] = []
                manager_allocations[mid].append(alloc.inspector_id)

            # Build inspector_salary_map: inspector_name -> monthly_salary
            inspector_salary_map = {}
            for sal in InspectorSalary.objects.all():
                inspector_salary_map[sal.inspector_name] = float(sal.monthly_salary)

            return _cors(JsonResponse({
                'success': True,
                'users': users_list,
                'inspector_users': inspector_users,
                'manager_allocations': manager_allocations,
                'inspector_salary_map': inspector_salary_map,
            }))
        except Exception as e:
            return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))

    # ── POST: handle actions ──
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')

            # ── add_user ──
            if action == 'add_user':
                username = data.get('username', '').strip()
                email = data.get('email', '').strip()
                password = data.get('password', '')
                password2 = data.get('password2', '')
                first_name = data.get('first_name', '').strip()
                last_name = data.get('last_name', '').strip()
                role = data.get('role', 'inspector')
                monthly_salary = data.get('monthly_salary')

                # Validation
                if not username or not email or not password or not first_name or not last_name:
                    return _cors(JsonResponse({'success': False, 'error': 'All fields are required (username, email, password, first_name, last_name).'}))

                if password != password2:
                    return _cors(JsonResponse({'success': False, 'error': 'Passwords do not match.'}))

                if User.objects.filter(username=username).exists():
                    return _cors(JsonResponse({'success': False, 'error': f'Username "{username}" is already taken.'}))

                if User.objects.filter(email=email).exists():
                    return _cors(JsonResponse({'success': False, 'error': f'Email "{email}" is already in use.'}))

                user = User.objects.create(
                    username=username,
                    email=email,
                    password=make_password(password),
                    first_name=first_name,
                    last_name=last_name,
                )
                user.role = role
                user.save()

                # Create InspectorSalary record if salary provided
                if monthly_salary is not None and monthly_salary != '':
                    try:
                        salary_val = float(monthly_salary)
                        inspector_name = f'{first_name} {last_name}'.upper()
                        InspectorSalary.objects.update_or_create(
                            inspector_name=inspector_name,
                            defaults={'monthly_salary': salary_val},
                        )
                    except (ValueError, TypeError):
                        pass

                return _cors(JsonResponse({'success': True, 'message': f'User "{username}" created successfully.', 'user_id': user.id}))

            # ── edit_user ──
            elif action == 'edit_user':
                user_id = data.get('user_id')
                if not user_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id.'}))

                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': f'User with id {user_id} not found.'}))

                email = data.get('email', '').strip()
                first_name = data.get('first_name', '').strip()
                last_name = data.get('last_name', '').strip()
                role = data.get('role', '')

                if email:
                    # Check uniqueness excluding current user
                    if User.objects.filter(email=email).exclude(id=user_id).exists():
                        return _cors(JsonResponse({'success': False, 'error': f'Email "{email}" is already in use by another user.'}))
                    user.email = email
                if first_name:
                    user.first_name = first_name
                if last_name:
                    user.last_name = last_name
                if role:
                    user.role = role

                user.save()

                # Handle inspector_manager allocated inspector ids
                allocated_inspector_ids = data.get('allocated_inspector_ids')
                if allocated_inspector_ids is not None and role == 'inspector_manager':
                    # Clear existing allocations for this manager
                    InspectorManagerAllocation.objects.filter(manager=user).delete()
                    # Create new allocations
                    for inspector_user_id in allocated_inspector_ids:
                        try:
                            inspector_user = User.objects.get(id=inspector_user_id)
                            InspectorManagerAllocation.objects.create(
                                manager=user,
                                inspector=inspector_user,
                            )
                        except User.DoesNotExist:
                            continue

                return _cors(JsonResponse({'success': True, 'message': f'User "{user.username}" updated successfully.'}))

            # ── toggle_user_status / activate_user / deactivate_user ──
            elif action in ('toggle_user_status', 'activate_user', 'deactivate_user'):
                user_id = data.get('user_id')
                if not user_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id.'}))

                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': f'User with id {user_id} not found.'}))

                if action == 'activate_user':
                    user.is_active = True
                elif action == 'deactivate_user':
                    user.is_active = False
                else:
                    user.is_active = not user.is_active
                user.save()

                status_str = 'activated' if user.is_active else 'deactivated'
                return _cors(JsonResponse({'success': True, 'message': f'User "{user.username}" {status_str}.', 'is_active': user.is_active}))

            # ── reset_password ──
            elif action == 'reset_password':
                user_id = data.get('user_id')
                new_password = data.get('new_password', '')

                if not user_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id.'}))
                if not new_password:
                    return _cors(JsonResponse({'success': False, 'error': 'New password is required.'}))

                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': f'User with id {user_id} not found.'}))

                user.password = make_password(new_password)
                user.save()

                return _cors(JsonResponse({'success': True, 'message': f'Password for "{user.username}" has been reset.'}))

            # ── delete_user (soft delete: deactivates account, preserves all data) ──
            elif action == 'delete_user':
                user_id = data.get('user_id')
                if not user_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id.'}))

                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': f'User with id {user_id} not found.'}))

                # Don't allow deactivating developer accounts
                if getattr(user, 'role', '') == 'developer':
                    return _cors(JsonResponse({'success': False, 'error': 'Cannot deactivate developer accounts.'}))

                username = user.username
                # Soft delete: deactivate account so they can't log in but all data
                # (system logs, sent_by, uploaded_by, approved_by, etc.) is preserved
                user.is_active = False
                user.save(update_fields=['is_active'])

                return _cors(JsonResponse({
                    'success': True,
                    'message': f'User "{username}" deactivated. All historical data preserved.'
                }))

            # ── reassign_and_delete (inspector departure) ──
            elif action == 'reassign_and_delete':
                user_id = data.get('user_id')
                reassign_to_id = data.get('reassign_to')
                if not user_id or not reassign_to_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id or reassign_to.'}))

                try:
                    departing = User.objects.get(id=user_id)
                    new_inspector = User.objects.get(id=reassign_to_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': 'User not found.'}))

                if getattr(departing, 'role', '') == 'developer':
                    return _cors(JsonResponse({'success': False, 'error': 'Cannot deactivate developer accounts.'}))

                departing_name = f"{departing.first_name} {departing.last_name}".strip() or departing.username
                new_name = f"{new_inspector.first_name} {new_inspector.last_name}".strip() or new_inspector.username

                # Reassign client allocations (ClientAllocation model if exists)
                from ..models import Client
                reassigned_clients = 0
                try:
                    from ..models import ClientAllocation
                    reassigned_clients = ClientAllocation.objects.filter(inspector_name__iexact=departing_name).update(inspector_name=new_name)
                    if not reassigned_clients:
                        reassigned_clients = ClientAllocation.objects.filter(inspector_name__iexact=departing.username).update(inspector_name=new_name)
                except Exception:
                    pass

                # Reassign InspectorMapping if exists
                try:
                    from ..models import InspectorMapping
                    InspectorMapping.objects.filter(inspector_name__iexact=departing_name).update(inspector_name=new_name)
                except Exception:
                    pass

                # Log the reassignment
                from ..models import SystemLog
                SystemLog.log_activity(
                    user=request.user if request.user.is_authenticated else None,
                    action='USER_MANAGEMENT',
                    page='user-management',
                    object_type='inspector_reassignment',
                    object_id=str(user_id),
                    description=f'Facilities reassigned from {departing_name} to {new_name}. {reassigned_clients} allocation(s) transferred. User "{departing.username}" deactivated.',
                )

                # Soft delete the departing inspector
                departing.is_active = False
                departing.save(update_fields=['is_active'])

                return _cors(JsonResponse({
                    'success': True,
                    'message': f'Facilities reassigned from {departing_name} to {new_name}. {reassigned_clients} allocation(s) transferred. User deactivated.'
                }))

            # ── send_reset_email ──
            elif action == 'send_reset_email':
                user_id = data.get('user_id')
                if not user_id:
                    return _cors(JsonResponse({'success': False, 'error': 'Missing user_id.'}))
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    return _cors(JsonResponse({'success': False, 'error': f'User with id {user_id} not found.'}))
                if not user.email:
                    return _cors(JsonResponse({'success': False, 'error': f'User "{user.username}" has no email address.'}))
                try:
                    from django.contrib.auth.tokens import default_token_generator
                    from django.utils.http import urlsafe_base64_encode
                    from django.utils.encoding import force_bytes
                    from django.core.mail import send_mail
                    from django.conf import settings as _settings
                    uid = urlsafe_base64_encode(force_bytes(user.pk))
                    token = default_token_generator.make_token(user)
                    reset_url = f'http://localhost:3000/reset-password/{uid}/{token}/'
                    # TEST MODE: redirect all emails to dev address
                    send_mail(
                        subject='Password Reset Request',
                        message=f'Hi {user.first_name or user.username},\n\nClick the link below to reset your password:\n{reset_url}\n\nIf you did not request this, please ignore this email.',
                        from_email=getattr(_settings, 'DEFAULT_FROM_EMAIL', ''),
                        recipient_list=['ethansevenster5@gmail.com'],
                    )
                    return _cors(JsonResponse({'success': True, 'message': f'Password reset email sent to {user.email}.'}))
                except Exception as e:
                    return _cors(JsonResponse({'success': False, 'error': f'Failed to send email: {e}'}))

            # ── send_kpi_emails ──
            elif action == 'send_kpi_emails':
                test_email = data.get('test_email', '').strip()
                try:
                    from main.services.kpi_email_service import send_kpi_email_to_inspector, send_kpi_emails_to_all, compute_inspector_kpi, build_kpi_email_html
                    from django.core.mail import EmailMessage
                    from django.conf import settings as _settings

                    if test_email:
                        # Send a single test KPI email to the specified address
                        inspector_users = User.objects.filter(role='inspector', is_active=True)
                        user_to_test = inspector_users.first()
                        if not user_to_test:
                            return _cors(JsonResponse({'success': False, 'error': 'No active inspectors found to generate test KPI data.'}))
                        name = user_to_test.get_full_name() or user_to_test.username
                        first_name = user_to_test.first_name or name.split()[0]
                        kpi_data = compute_inspector_kpi(name)
                        html_body = build_kpi_email_html(kpi_data, first_name)
                        subject = f"[TEST] {kpi_data['quarter_label']} KPI Report — {name}"
                        msg = EmailMessage(
                            subject=subject, body=html_body,
                            from_email=getattr(_settings, 'DEFAULT_FROM_EMAIL', ''),
                            to=[test_email],
                        )
                        msg.content_subtype = 'html'
                        msg.send()
                        return _cors(JsonResponse({'success': True, 'message': f'Test KPI email sent to {test_email} using data for {name}.', 'sent': 1, 'failed': 0}))
                    else:
                        results = send_kpi_emails_to_all()
                        sent = sum(1 for r in results if r['success'])
                        failed = sum(1 for r in results if not r['success'])
                        return _cors(JsonResponse({'success': True, 'message': f'KPI emails sent: {sent} succeeded, {failed} failed.', 'sent': sent, 'failed': failed, 'results': results}))
                except Exception as e:
                    return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))

            else:
                return _cors(JsonResponse({'success': False, 'error': f'Unknown action: {action}'}))

        except json.JSONDecodeError:
            return _cors(JsonResponse({'success': False, 'error': 'Invalid JSON in request body.'}, status=400))
        except Exception as e:
            return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))

    return _cors(JsonResponse({'success': False, 'error': 'Method not allowed.'}, status=405))


from django.views.decorators.csrf import csrf_exempt as _csrf_exempt


@_csrf_exempt
def api_home_stats(request):
    """Lightweight endpoint returning total client and inspection counts for the home page."""
    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        from ..models import FoodSafetyAgencyInspection as _I, Client as _C, SystemLog as _SL
        clients_count = _C.objects.count()
        inspections_count = _I.objects.count()

        # Recent activity from SystemLog
        try:
            logs = list(
                _SL.objects.select_related('user')
                .order_by('-timestamp')[:5]
                .values('action', 'description', 'timestamp', 'user__username')
            )
            recent_activities = [
                {
                    'action': l['action'],
                    'description': l['description'] or '',
                    'username': l['user__username'] or 'System',
                    'timestamp': l['timestamp'].isoformat() if l['timestamp'] else '',
                }
                for l in logs
            ]
        except Exception:
            recent_activities = []

        return _cors(JsonResponse({
            'clients': clients_count,
            'inspections': inspections_count,
            'recent_activities': recent_activities,
        }))
    except Exception as e:
        return _cors(JsonResponse({'clients': 0, 'inspections': 0, 'recent_activities': [], 'error': str(e)}))


@_csrf_exempt
def api_lab_analytics(request):
    """Lab analytics dashboard data for lab technician role."""
    from ..models import FoodSafetyAgencyInspection as _I
    from django.db.models import Count, Q

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        base = _I.objects.filter(is_sample_taken=True)

        _LAB_DISPLAY = {
            'lab_a': 'Food Safety Laboratory',
            'lab_b': 'Merieux NutriSciences',
            'lab_c': 'AGRI Food Laboratory (SGS)',
            'lab_d': 'SANBI',
            'lab_e': 'SMT',
            'lab_f': 'ARC',
        }

        # Apply date filters from query params
        _date_from = request.GET.get('date_from')
        _date_to = request.GET.get('date_to')
        _lab_filter = request.GET.get('lab')
        _commodity_filter = request.GET.get('commodity')
        if _date_from:
            base = base.filter(date_of_inspection__gte=_date_from)
        if _date_to:
            base = base.filter(date_of_inspection__lte=_date_to)
        if _lab_filter:
            # Support both raw key (lab_a) and display name (Food Safety Laboratory)
            _LAB_REVERSE = {v: k for k, v in _LAB_DISPLAY.items()}
            _lab_key = _LAB_REVERSE.get(_lab_filter, _lab_filter)
            base = base.filter(lab=_lab_key)
        if _commodity_filter:
            base = base.filter(commodity=_commodity_filter)

        # Summary stats
        total_samples   = base.count()
        _insp_qs = _I.objects.all()
        if _date_from:
            _insp_qs = _insp_qs.filter(date_of_inspection__gte=_date_from)
        if _date_to:
            _insp_qs = _insp_qs.filter(date_of_inspection__lte=_date_to)
        total_inspections = _insp_qs.count()
        needs_coa       = base.filter(coa_uploaded_date__isnull=True).count()
        needs_retest    = base.filter(needs_retest__in=['Yes', 'YES', 'yes']).count()
        fat_count       = base.filter(fat=True).count()
        protein_count   = base.filter(protein=True).count()
        calcium_count   = base.filter(calcium=True).count()
        dna_count       = base.filter(dna=True).count()
        total_tests     = fat_count + protein_count + calcium_count + dna_count

        # By lab (filtered for current view)
        labs = [
            {'lab': _LAB_DISPLAY.get(row['lab'], row['lab']), 'n': row['n']}
            for row in base.exclude(lab='').exclude(lab__isnull=True)
                .values('lab').annotate(n=Count('id')).order_by('-n')[:10]
        ]

        # By lab (ALL labs, unfiltered — always show full picture, include 0-count labs)
        _all_samples = _I.objects.filter(is_sample_taken=True)
        _lab_counts = {row['lab']: row['n'] for row in _all_samples.exclude(lab='').exclude(lab__isnull=True).values('lab').annotate(n=Count('id'))}
        all_labs_stats = [
            {'lab': _LAB_DISPLAY.get(key, key), 'n': _lab_counts.get(key, 0)}
            for key in _LAB_DISPLAY.keys()
        ]
        all_labs_stats.sort(key=lambda x: -x['n'])

        # By commodity (filtered)
        commodities = list(
            base.values('commodity').annotate(n=Count('id')).order_by('-n')
        )

        # By commodity (ALL, unfiltered)
        all_commodities_stats = list(
            _all_samples.exclude(commodity='').exclude(commodity__isnull=True)
            .values('commodity').annotate(n=Count('id')).order_by('-n')
        )

        # Monthly trend — last 6 months
        import datetime
        from django.utils import timezone as _tz
        today = _tz.now().date()
        monthly = []
        for i in range(5, -1, -1):
            # First day of the month i months ago
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            first = datetime.date(y, m, 1)
            if m == 12:
                last = datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
            else:
                last = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
            count = base.filter(date_of_inspection__gte=first, date_of_inspection__lte=last).count()
            monthly.append({'month': first.strftime('%b %Y'), 'count': count})

        # Recent samples
        recent_qs = base.order_by('-date_of_inspection')[:50].values(
            'client_name', 'product_name', 'commodity', 'lab',
            'needs_retest', 'fat', 'protein', 'calcium', 'dna', 'date_of_inspection'
        )
        recent = []
        for r in recent_qs:
            tests = [t for t in ['fat', 'protein', 'calcium', 'dna'] if r.get(t)]
            recent.append({
                'client_name':    r['client_name'] or '',
                'product_name':   r['product_name'] or '',
                'commodity':      r['commodity'] or '',
                'lab':            _LAB_DISPLAY.get(r['lab'], r['lab']) if r.get('lab') else '',
                'needs_retest':   r['needs_retest'] or 'NO',
                'tests':          tests,
                'date':           r['date_of_inspection'].isoformat() if r['date_of_inspection'] else '',
            })

        # Filter options — always show ALL known labs (even with 0 samples)
        all_labs_options = list(_LAB_DISPLAY.values())
        _all_base = _I.objects.filter(is_sample_taken=True)
        all_commodities_options = list(_all_base.exclude(commodity='').exclude(commodity__isnull=True).values_list('commodity', flat=True).distinct().order_by('commodity'))

        return _cors(JsonResponse({
            'success': True,
            'total_samples':      total_samples,
            'total_inspections':  total_inspections,
            'needs_coa':          needs_coa,
            'needs_retest':       needs_retest,
            'total_tests':     total_tests,
            'fat_count':       fat_count,
            'protein_count':   protein_count,
            'calcium_count':   calcium_count,
            'dna_count':       dna_count,
            'labs':            labs,
            'commodities':     commodities,
            'allLabsStats':    all_labs_stats,
            'allCommoditiesStats': all_commodities_stats,
            'monthly':         monthly,
            'recent':          recent,
            'allLabs':         sorted(set(all_labs_options)),
            'allCommodities':  sorted(set(all_commodities_options)),
        }))

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_me(request):
    """Return the currently authenticated user's profile and role."""
    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        r['Access-Control-Allow-Credentials'] = 'true'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if not request.user.is_authenticated:
        return _cors(JsonResponse({'authenticated': False}, status=401))

    user = request.user
    return _cors(JsonResponse({
        'authenticated': True,
        'id': user.id,
        'username': user.username,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'email': user.email,
        'role': getattr(user, 'role', 'inspector') or 'inspector',
    }))


@_csrf_exempt
def api_login(request):
    """CSRF-exempt JSON login endpoint for the Next.js frontend."""
    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With'
        r['Access-Control-Allow-Credentials'] = 'true'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed.'}, status=405))

    try:
        content_type = request.content_type or ''
        if 'application/json' in content_type:
            import json as _json
            body = _json.loads(request.body)
            username = body.get('username', '').strip()
            password = body.get('password', '').strip()
        else:
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()

        if not username or not password:
            return _cors(JsonResponse({'success': False, 'error': 'Username and password are required.'}, status=400))

        # Case-insensitive username lookup
        from django.contrib.auth import get_user_model, authenticate, login as _login
        _User = get_user_model()
        try:
            actual_user = _User.objects.get(username__iexact=username)
            username = actual_user.username
        except _User.DoesNotExist:
            pass

        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            _login(request, user)
            from django.utils import timezone
            request.session['last_activity'] = timezone.now().isoformat()
            request.session['authenticated'] = True
            request.session.modified = True
            return _cors(JsonResponse({'success': True, 'username': user.username}))
        else:
            return _cors(JsonResponse({'success': False, 'error': 'Invalid username or password.'}, status=401))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_get_inspection_files(request):
    """API endpoint for the Next.js frontend to get inspection files (no login_required, no CSRF).
    Replicates the logic of get_inspection_files in core_views.py."""

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))

    try:
        import json as _json
        from .core_views import get_inspection_files_local

        data = _json.loads(request.body)
        group_id = data.get('group_id', '')
        client_name = data.get('client_name', '')
        inspection_date = data.get('inspection_date', '')

        # Strip "btn-" prefix from client name if present
        if client_name and client_name.startswith('btn-'):
            client_name = client_name[4:]

        # Clean Unicode escapes from inspection_date
        if isinstance(inspection_date, str):
            inspection_date = inspection_date.replace('\\u002D', '-')
            inspection_date = inspection_date.replace('\\u002F', '/')
            inspection_date = inspection_date.replace('\\u0020', ' ')

        local_files = get_inspection_files_local(client_name, inspection_date, False)

        has_files = local_files and any(file_list for file_list in local_files.values() if file_list)

        if not has_files:
            empty_files = {
                'rfi': [], 'invoice': [], 'lab': [], 'lab_form': [],
                'retest': [], 'compliance': [], 'occurrence': [],
                'composition': [], 'other': []
            }
            resp = JsonResponse({
                'success': True, 'files': empty_files,
                'client_name': client_name, 'inspection_date': inspection_date,
            })
        else:
            resp = JsonResponse({
                'success': True, 'files': local_files,
                'client_name': client_name, 'inspection_date': inspection_date,
            })

        return _cors(resp)

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}))


@_csrf_exempt
def api_serve_file(request):
    """Serve an inspection file without login_required (for Next.js frontend)."""
    from django.http import FileResponse, HttpResponse, Http404
    from django.conf import settings as _settings
    import os, mimetypes, urllib.parse

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(HttpResponse())

    relative_path = request.GET.get('file', '')
    action = request.GET.get('action', 'view')

    if not relative_path:
        raise Http404('No file specified')

    # Prevent directory traversal
    safe_path = os.path.normpath(relative_path).lstrip('/\\')
    file_path = os.path.join(_settings.MEDIA_ROOT, safe_path)
    file_path = os.path.normpath(file_path)
    if not file_path.startswith(os.path.normpath(_settings.MEDIA_ROOT)):
        raise Http404('Invalid file path')

    if not os.path.isfile(file_path):
        raise Http404(f'File not found: {safe_path}')

    mime_type, _ = mimetypes.guess_type(file_path)
    mime_type = mime_type or 'application/octet-stream'

    f = open(file_path, 'rb')
    response = FileResponse(f, content_type=mime_type)
    if action == 'download':
        filename = os.path.basename(file_path)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        response['Content-Disposition'] = 'inline'
    return _cors(response)


@_csrf_exempt
def api_delete_file(request):
    """Delete an inspection file without login_required (for Next.js frontend)."""
    from django.http import HttpResponse
    from django.conf import settings as _settings
    import os

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, DELETE, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method not in ('POST', 'DELETE'):
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))

    try:
        import json as _json
        body = _json.loads(request.body)
        relative_path = body.get('file', '')
    except Exception:
        relative_path = request.GET.get('file', '')

    if not relative_path:
        return _cors(JsonResponse({'success': False, 'error': 'No file specified'}, status=400))

    safe_path = os.path.normpath(relative_path).lstrip('/\\')
    file_path = os.path.join(_settings.MEDIA_ROOT, safe_path)
    file_path = os.path.normpath(file_path)
    if not file_path.startswith(os.path.normpath(_settings.MEDIA_ROOT)):
        return _cors(JsonResponse({'success': False, 'error': 'Invalid file path'}, status=400))

    if not os.path.isfile(file_path):
        return _cors(JsonResponse({'success': False, 'error': 'File not found'}, status=404))

    os.remove(file_path)
    return _cors(JsonResponse({'success': True, 'message': f'File deleted: {os.path.basename(file_path)}'}))


@_csrf_exempt
def api_upload_document(request):
    """API upload endpoint for the Next.js frontend (no login_required, no CSRF).
    Proxies to the same logic as upload_document in core_views.py."""

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))

    # Delegate to the existing upload_document view in core_views
    # but bypass authentication check
    from .core_views import upload_document as _core_upload

    # Use a real user from the DB so FK constraints on uploaded_by fields are satisfied
    from django.contrib.auth import get_user_model as _get_user_model
    _User = _get_user_model()
    _real_user = (
        _User.objects.filter(is_superuser=True).first()
        or _User.objects.filter(is_staff=True).first()
        or _User.objects.first()
    )

    original_user = request.user
    if _real_user:
        request.user = _real_user

    try:
        response = _core_upload(request)
    finally:
        request.user = original_user

    return _cors(response)


@_csrf_exempt
def api_inspection_form_data(request):
    """API endpoint returning dropdown options for the Add Inspection form."""
    from ..models import FoodSafetyAgencyInspection as _Insp, Client as _Client

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        from django.core.cache import cache
        cache_key = 'add_inspection_form_data'
        cached = cache.get(cache_key)

        if cached:
            clients_with_towns = cached['clients']
            towns_list = cached['towns']
            all_groups = cached['groups']
        else:
            town_lookup = dict(
                _Insp.objects.exclude(town__isnull=True).exclude(town='')
                .order_by('client_name', '-date_of_inspection')
                .distinct('client_name').values_list('client_name', 'town')
            )
            clients_with_towns = [
                {
                    'name': c.name,
                    'email': c.email or '',
                    'town': town_lookup.get(c.name, ''),
                    'corporate_group': c.corporate_group or '',
                    'group_type': c.group_type or '',
                    'facility_type': c.facility_type or '',
                }
                for c in _Client.objects.all().order_by('name')
            ]
            towns_list = list(
                _Insp.objects.exclude(town__isnull=True).exclude(town='')
                .values_list('town', flat=True).distinct().order_by('town')
            )
            default_corporate_groups = [
                'Pick n Pay - Franchise', 'Pick n Pay - Corporate', 'Fruit & Veg', 'OK Foods',
                'Checkers', 'Spar', 'SuperSpar', 'Spar - Northrand', 'Shoprite', 'Massmart',
                'Chester Butcheries', 'Boxer', 'Food Lovers Market', 'Cambridge', 'Woolworths',
                'Jwayelani', 'Usave',
            ]
            db_groups = list(
                _Insp.objects.exclude(corporate_group__isnull=True).exclude(corporate_group='')
                .exclude(corporate_group__in=['Not Applicable', 'Other'])
                .values_list('corporate_group', flat=True).distinct()
            )
            all_groups = list(default_corporate_groups)
            extra_groups = sorted(set(db_groups) - set(default_corporate_groups))
            all_groups.extend(extra_groups)
            cache.set(cache_key, {'clients': clients_with_towns, 'towns': towns_list, 'groups': all_groups}, 600)

        # Get inspectors
        inspectors = list(
            _Insp.objects.exclude(inspector_name__isnull=True).exclude(inspector_name='')
            .values_list('inspector_name', flat=True).distinct().order_by('inspector_name')
        )

        return _cors(JsonResponse({
            'success': True,
            'clients': clients_with_towns,
            'towns': towns_list,
            'corporate_groups': all_groups,
            'inspectors': inspectors,
        }))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_add_inspection(request):
    """API endpoint to create a new inspection group with products."""
    import json as _json
    from ..models import (
        FoodSafetyAgencyInspection as _Insp,
        InspectionGroup as _Group,
        Client as _Client,
    )
    from django.db.models import Min

    def _cors(response):
        response['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        response['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'
        return response

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))

    try:
        data = _json.loads(request.body)

        # Validate required fields
        missing = []
        if not data.get('client_name', '').strip():
            missing.append('Client Name')
        if not data.get('town', '').strip():
            missing.append('Town')
        if not data.get('corporate_group', '').strip():
            missing.append('Corporate Group')
        if not data.get('group_type', '').strip():
            missing.append('Group Type')
        if not data.get('facility_type', '').strip():
            missing.append('Facility Type')
        if not data.get('date_of_inspection', '').strip():
            missing.append('Date of Inspection')
        if missing:
            return _cors(JsonResponse({'success': False, 'error': f"Required fields missing: {', '.join(missing)}"}))

        is_occurrence = bool(data.get('is_occurrence_report', False))
        products_data = data.get('products', [])
        if not is_occurrence:
            if not products_data:
                return _cors(JsonResponse({'success': False, 'error': 'Please add at least one product'}))

            for i, p in enumerate(products_data):
                if not p.get('product_name', '').strip():
                    return _cors(JsonResponse({'success': False, 'error': f'Product name is required for product {i + 1}'}))

        from django.utils.dateparse import parse_date as _pd
        date_obj = _pd(data['date_of_inspection'])
        if not date_obj:
            return _cors(JsonResponse({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}))

        # Resolve inspector name: payload > authenticated user > fallback
        inspector_name = data.get('inspector_name', '').strip()
        if not inspector_name or inspector_name == 'API User':
            if hasattr(request, 'user') and request.user.is_authenticated:
                inspector_name = f"{request.user.first_name} {request.user.last_name}".strip()
                if not inspector_name:
                    inspector_name = request.user.username
            else:
                inspector_name = 'API User'
        data['inspector_name'] = inspector_name

        with transaction.atomic():
            # Get or create client
            client_name = data['client_name'].strip()
            client = _Client.objects.filter(name__iexact=client_name).first()
            if not client:
                client = _Client.objects.create(
                    name=client_name,
                    town=data.get('town', ''),
                    corporate_group=data.get('corporate_group', ''),
                    group_type=data.get('group_type', ''),
                    facility_type=data.get('facility_type', ''),
                )

            # Create parent InspectionGroup
            parent_group = _Group.objects.create(
                client=client,
                client_name=client_name,
                date_of_inspection=date_obj,
                inspector_name=data.get('inspector_name', 'API User'),
                town=data.get('town', ''),
                facility_type=data.get('facility_type', ''),
                group_type=data.get('group_type', ''),
                corporate_group=data.get('corporate_group', ''),
                additional_email=data.get('additional_email', ''),
                comment=data.get('comment', ''),
                km_traveled=float(data.get('km_traveled', 0) or 0),
                hours=float(data.get('hours', 0) or 0),
                is_manual=True,
            )

            # Generate account code
            from .core_views import generate_unique_internal_account_code
            account_code = generate_unique_internal_account_code(
                client_name=client_name,
                date_of_inspection=date_obj,
                facility_type=data.get('facility_type', ''),
                group_type=data.get('group_type', ''),
                commodity='',
                corporate_group=data.get('corporate_group', ''),
                client_id=client.id,
            )

            total_created = 0

            # Occurrence report fields (shared across all child inspections)
            occurrence_kwargs = {}
            if is_occurrence:
                occurrence_kwargs = {
                    'is_occurrence_report': True,
                    'registration_code': data.get('registration_code', ''),
                    'physical_address': data.get('physical_address', ''),
                    'telephone': data.get('telephone', ''),
                }
                time_of_visit_str = data.get('time_of_visit', '')
                if time_of_visit_str:
                    from django.utils.dateparse import parse_time as _pt
                    occurrence_kwargs['time_of_visit'] = _pt(time_of_visit_str)

            if is_occurrence and not products_data:
                # Create a single inspection record for occurrence reports with no products
                min_rid = _Insp.objects.filter(is_manual=True).aggregate(Min('remote_id'))['remote_id__min']
                remote_id = -1 if (min_rid is None or min_rid >= 0) else min_rid - 1

                _Insp.objects.create(
                    inspection_group=parent_group,
                    client=client,
                    client_name=client_name,
                    date_of_inspection=date_obj,
                    commodity='',
                    product_name='Occurrence Report',
                    product_class='',
                    inspector_name=data.get('inspector_name', 'API User'),
                    town=data.get('town', ''),
                    is_sample_taken=False,
                    needs_retest='NO',
                    km_traveled=float(data.get('km_traveled', 0) or 0),
                    hours=float(data.get('hours', 0) or 0),
                    additional_email=data.get('additional_email', ''),
                    comment=data.get('comment', ''),
                    is_manual=True,
                    inspected=True,
                    follow_up=bool(data.get('follow_up', False)),
                    dispensation_application=bool(data.get('dispensation_application', False)),
                    corporate_group=data.get('corporate_group', ''),
                    group_type=data.get('group_type', ''),
                    facility_type=data.get('facility_type', ''),
                    remote_id=remote_id,
                    inspection_sequence=1,
                    internal_account_code=account_code,
                    **occurrence_kwargs,
                )
                total_created = 1
            else:
                for seq, prod in enumerate(products_data, start=1):
                    # Generate unique negative remote_id
                    min_rid = _Insp.objects.filter(is_manual=True).aggregate(Min('remote_id'))['remote_id__min']
                    remote_id = -1 if (min_rid is None or min_rid >= 0) else min_rid - 1

                    _Insp.objects.create(
                        inspection_group=parent_group,
                        client=client,
                        client_name=client_name,
                        date_of_inspection=date_obj,
                        commodity=prod.get('commodity', ''),
                        product_name=prod.get('product_name', ''),
                        product_class=prod.get('product_class', ''),
                        inspector_name=data.get('inspector_name', 'API User'),
                        town=data.get('town', ''),
                        is_sample_taken=bool(prod.get('is_sample_taken', False)),
                        needs_retest=prod.get('needs_retest', 'NO'),
                        fat=bool(prod.get('fat', False)),
                        protein=bool(prod.get('protein', False)),
                        calcium=bool(prod.get('calcium', False)),
                        dna=bool(prod.get('dna', False)),
                        lab=prod.get('lab', ''),
                        km_traveled=float(data.get('km_traveled', 0) or 0),
                        hours=float(data.get('hours', 0) or 0),
                        additional_email=data.get('additional_email', ''),
                        comment=data.get('comment', ''),
                        is_manual=True,
                        inspected=True,
                        follow_up=bool(data.get('follow_up', False)),
                        dispensation_application=bool(data.get('dispensation_application', False)),
                        corporate_group=data.get('corporate_group', ''),
                        group_type=data.get('group_type', ''),
                        facility_type=data.get('facility_type', ''),
                        remote_id=remote_id,
                        inspection_sequence=seq,
                        internal_account_code=account_code,
                        **occurrence_kwargs,
                    )
                    total_created += 1

            # Clear cache
            from django.core.cache import cache as _fc
            _fc.delete('add_inspection_form_data')

            if is_occurrence:
                commodity_summary = 'Occurrence Report'
            else:
                from collections import Counter
                commodity_counts = Counter(p.get('commodity', '') for p in products_data)
                commodity_summary = ', '.join([f"{c}: {n}" for c, n in commodity_counts.items()])

        return _cors(JsonResponse({
            'success': True,
            'message': f"Created {total_created} inspection(s) for {client_name} ({commodity_summary})",
            'group_id': parent_group.id,
            'total_created': total_created,
        }))

    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: System Logs
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_system_logs(request):
    from ..models import SystemLog

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        from datetime import datetime as _dt
        import math as _math
        # Exclude duplicate/ghost logs from v4 Django middleware sharing the same DB
        # Filter out: generic middleware entries, and entries from users who only exist on v4
        # (Simphiwe's file uploads are ghost entries from shared session DB)
        logs = SystemLog.objects.select_related('user').order_by('-timestamp').exclude(
            description__icontains='uploaded a file on'
        ).exclude(
            description__icontains='went to'
        ).exclude(
            description__icontains='viewed', page__startswith='/api/'
        ).exclude(
            # Exclude file uploads attributed to v4-only users via shared sessions
            user__username='Simphiwe', page__startswith='/api/'
        ).exclude(
            user__username='Armand', page__startswith='/api/'
        )
        user_filter = request.GET.get('user', '')
        action_filter = request.GET.get('action', '')
        page_filter = request.GET.get('page_filter', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        show_all = request.GET.get('show_all', 'false') == 'true'
        try:
            page_num = max(1, int(request.GET.get('page_num', 1)))
        except (ValueError, TypeError):
            page_num = 1

        if user_filter:
            logs = logs.filter(user__username__icontains=user_filter)
        if action_filter:
            logs = logs.filter(action=action_filter)
        if page_filter:
            logs = logs.filter(page__icontains=page_filter)
        if date_from:
            try:
                logs = logs.filter(timestamp__date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if date_to:
            try:
                logs = logs.filter(timestamp__date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        total = logs.count()
        page_size = 50
        if show_all:
            logs_qs = logs[:1000]
            total_pages = 1
        else:
            offset = (page_num - 1) * page_size
            logs_qs = logs[offset:offset + page_size]
            total_pages = max(1, _math.ceil(total / page_size))

        logs_list = list(logs_qs)
        unique_ips = list({l.ip_address for l in logs_list if l.ip_address})
        geo_map = _resolve_ips_to_locations(unique_ips)

        results = [{
            'id': l.id,
            'timestamp': l.timestamp.isoformat() if l.timestamp else '',
            'username': (l.user.get_full_name() or l.user.username) if l.user else 'System',
            'action': l.action or '',
            'page': l.page or '',
            'description': l.description or '',
            'ip_address': l.ip_address or '',
            'location': geo_map.get(l.ip_address or '', {}).get('display', ''),
            'suburb': geo_map.get(l.ip_address or '', {}).get('suburb', ''),
            'city': geo_map.get(l.ip_address or '', {}).get('city', ''),
            'country_code': geo_map.get(l.ip_address or '', {}).get('country_code', ''),
        } for l in logs_list]

        # Dropdown options
        all_users = list(SystemLog.objects.select_related('user').values_list('user__username', flat=True).distinct().order_by('user__username'))
        all_users = [u for u in all_users if u]
        all_pages = list(SystemLog.objects.exclude(page__isnull=True).exclude(page='').values_list('page', flat=True).distinct().order_by('page'))

        from ..models import FoodSafetyAgencyInspection as _I
        from django.db.models import Count, Min, Max
        dup_pairs = list(
            _I.objects.values('client_name', 'date_of_inspection', 'inspector_name')
            .annotate(gc=Count('inspection_group', distinct=True))
            .filter(gc__gt=1).order_by('-date_of_inspection')[:100]
        )
        duplicates = []
        for dp in dup_pairs:
            ids = list(_I.objects.filter(
                client_name=dp['client_name'],
                date_of_inspection=dp['date_of_inspection'],
                inspector_name=dp['inspector_name'],
            ).values_list('id', flat=True).order_by('id'))
            duplicates.append({
                'client_name': dp['client_name'] or '',
                'inspector': dp['inspector_name'] or '',
                'date': dp['date_of_inspection'].isoformat() if dp['date_of_inspection'] else '',
                'count': dp['gc'],
                'first_id': ids[0] if ids else None,
                'last_id': ids[-1] if ids else None,
            })

        # Edit history
        edit_history = []
        edit_history_total = 0
        try:
            from ..models import InspectionEditHistory
            eh_qs = InspectionEditHistory.objects.select_related('edited_by', 'inspection_group').order_by('-edited_at')
            if user_filter:
                eh_qs = eh_qs.filter(edited_by__username__icontains=user_filter)
            if date_from:
                try:
                    eh_qs = eh_qs.filter(edited_at__date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
                except ValueError:
                    pass
            if date_to:
                try:
                    eh_qs = eh_qs.filter(edited_at__date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
                except ValueError:
                    pass
            edit_history_total = eh_qs.count()
            edit_history = [{
                'id': eh.id,
                'edited_at': eh.edited_at.isoformat() if eh.edited_at else '',
                'edited_by': (eh.edited_by.get_full_name() or eh.edited_by.username) if eh.edited_by else 'System',
                'object_type': getattr(eh, 'object_type', ''),
                'client_name': getattr(eh, 'client_name', '') or '',
                'date_of_inspection': eh.date_of_inspection.isoformat() if getattr(eh, 'date_of_inspection', None) else '',
                'change_count': getattr(eh, 'change_count', len(eh.changes) if hasattr(eh, 'changes') else 0),
                'changes': eh.changes if hasattr(eh, 'changes') else {},
            } for eh in eh_qs[:200]]
        except Exception:
            pass

        return _cors(JsonResponse({
            'success': True,
            'total': total,
            'total_pages': total_pages,
            'page_num': page_num,
            'logs': results,
            'duplicates': duplicates,
            'duplicate_count': len(duplicates),
            'edit_history': edit_history,
            'edit_history_total': edit_history_total,
            'all_users': all_users,
            'all_pages': all_pages,
        }))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Export Sheet
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_export_sheet(request):
    from ..models import FoodSafetyAgencyInspection as _I
    from datetime import datetime as _dt, timedelta
    from collections import defaultdict
    from .core_views import generate_visit_hours_km_items, generate_test_line_items

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        today = _dt.now().date()
        yesterday = today - timedelta(days=1)
        date_from = request.GET.get('date_from', yesterday.strftime('%Y-%m-%d'))
        date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
        try:
            start_date = _dt.strptime(date_from, '%Y-%m-%d').date()
        except Exception:
            start_date = yesterday
        try:
            end_date = _dt.strptime(date_to, '%Y-%m-%d').date()
        except Exception:
            end_date = today

        inspections = _I.objects.filter(
            commodity__in=['RAW', 'PMP'], hours__isnull=False, km_traveled__isnull=False,
            date_of_inspection__gte=start_date, date_of_inspection__lte=end_date
        ).only(
            'id', 'commodity', 'date_of_inspection', 'inspector_name', 'client_name',
            'town', 'product_name', 'product_class', 'hours', 'km_traveled',
            'is_sample_taken', 'bought_sample', 'fat', 'protein', 'calcium', 'dna',
            'lab', 'invoice_number', 'corporate_group',
        ).order_by('client_name', 'date_of_inspection', 'commodity')

        visits = defaultdict(list)
        for insp in inspections:
            key = (insp.inspector_name or '', insp.client_name or '', str(insp.date_of_inspection) if insp.date_of_inspection else '')
            visits[key].append(insp)

        invoice_items = []
        counter = {}
        processed = 0

        for (inspector_name, client_name, _), vinsp in visits.items():
            vinsp.sort(key=lambda x: (x.commodity or '', x.product_name or ''))
            first = vinsp[0]
            processed += 1
            ikey = inspector_name or 'Unknown'
            counter[ikey] = counter.get(ikey, 0) + 1
            iid = counter[ikey]
            icode = ''.join([p[0].upper() for p in inspector_name.split() if p]) if inspector_name else ''
            dfmt = first.date_of_inspection.strftime('%y%m%d') if first.date_of_inspection else ''
            inv_ref = f"FSA-INV-{icode}-{dfmt}" if dfmt else ''
            rfi_ref = f"FSA-RFI-{icode}-{dfmt}" if dfmt else ''
            city = first.town or ''
            lab_name = 'Food Safety Laboratory' if first.lab else ''
            th = float(first.hours) if first.hours else 0
            tk = float(first.km_traveled) if first.km_traveled else 0

            if th > 0 or tk > 0:
                pmp = [i for i in vinsp if 'PMP' in (i.commodity or '').upper()]
                raw = [i for i in vinsp if 'RAW' in (i.commodity or '').upper()]
                if pmp and raw:
                    invoice_items.extend(generate_visit_hours_km_items(iid, pmp[0], inv_ref, rfi_ref, 'PMP', city, lab_name, th / 2, tk / 2))
                    invoice_items.extend(generate_visit_hours_km_items(iid, raw[0], inv_ref, rfi_ref, 'RAW', city, lab_name, th / 2, tk / 2))
                else:
                    ptype = 'RAW' if first.commodity and 'RAW' in first.commodity.upper() else 'PMP'
                    invoice_items.extend(generate_visit_hours_km_items(iid, first, inv_ref, rfi_ref, ptype, city, lab_name, th, tk))

            pmp = [i for i in vinsp if 'PMP' in (i.commodity or '').upper()]
            raw = [i for i in vinsp if 'RAW' in (i.commodity or '').upper()]
            if pmp:
                invoice_items.extend(generate_test_line_items(iid, pmp[0], inv_ref, rfi_ref, 'PMP', city, lab_name,
                    {'fat': any(i.fat and i.is_sample_taken for i in pmp), 'protein': any(i.protein and i.is_sample_taken for i in pmp), 'calcium': any(i.calcium and i.is_sample_taken for i in pmp)}))
            if raw:
                invoice_items.extend(generate_test_line_items(iid, raw[0], inv_ref, rfi_ref, 'RAW', city, lab_name,
                    {'fat': any(i.fat and i.is_sample_taken for i in raw), 'protein': any(i.protein and i.is_sample_taken for i in raw), 'dna': any(i.dna and i.is_sample_taken for i in raw)}))

        invoice_items.sort(key=lambda x: (x.get('client_name', ''), x.get('invoice_date', ''), x.get('item_code', '')))
        return _cors(JsonResponse({'success': True, 'items': invoice_items, 'total_items': len(invoice_items), 'inspections_processed': processed, 'date_from': start_date.strftime('%Y-%m-%d'), 'date_to': end_date.strftime('%Y-%m-%d')}))
    except Exception as e:
        import traceback; traceback.print_exc()
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Inspection Fees (unauthenticated proxy for React)
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_react_fees_get(request):
    """Return all inspection fees — no auth required for Next.js.

    Optional query param ?as_of_date=YYYY-MM-DD returns the historical
    rate that was active on that specific date instead of the current rate.
    """
    from ..models import InspectionFee
    from datetime import date as _date_cls

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        as_of_str = request.GET.get('as_of_date', '').strip()
        as_of_date = None
        if as_of_str:
            try:
                as_of_date = _date_cls.fromisoformat(as_of_str)
            except (ValueError, TypeError):
                as_of_date = None

        fees = InspectionFee.objects.all()
        fees_data = []
        for fee in fees:
            latest = fee.history.order_by('-effective_date').first()

            # If as_of_date provided, return the rate active on that date
            if as_of_date:
                # Find the latest history entry that was active on or before the requested date
                active_entry = fee.history.filter(effective_date__lte=as_of_date).order_by('-effective_date').first()
                if active_entry:
                    active_rate = float(active_entry.rate)
                    effective_date_for_display = active_entry.effective_date.isoformat()
                else:
                    # No historical entry exists before this date — fee was not yet defined
                    active_rate = 0.0
                    effective_date_for_display = None
            else:
                active_rate = float(fee.rate)
                effective_date_for_display = latest.effective_date.isoformat() if latest else None

            fees_data.append({
                'id': fee.id,
                'fee_code': fee.fee_code,
                'fee_name': fee.fee_name,
                'rate': active_rate,
                'current_rate': float(fee.rate),
                'description': fee.description,
                'last_updated': fee.last_updated.isoformat() if fee.last_updated else None,
                'effective_date': effective_date_for_display,
                'has_history': fee.history.exists(),
                'history_count': fee.history.count(),
            })
        return _cors(JsonResponse({
            'fees': fees_data,
            'as_of_date': as_of_date.isoformat() if as_of_date else None,
        }))
    except Exception as e:
        return _cors(JsonResponse({'error': str(e)}, status=500))


@_csrf_exempt
def api_react_fees_update(request):
    """Update inspection fees — no auth required for Next.js (POST only)."""
    import json
    from ..models import InspectionFee, FeeHistory
    from datetime import date
    from decimal import Decimal

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _cors(JsonResponse({'error': 'POST required'}, status=405))

    try:
        data = json.loads(request.body)
        fees_data = data.get('fees', [])

        # Parse effective_date from payload (default to today)
        eff_date_str = data.get('effective_date', '')
        try:
            effective_date = date.fromisoformat(eff_date_str) if eff_date_str else date.today()
        except (ValueError, TypeError):
            effective_date = date.today()

        notes = data.get('notes', '')
        is_future = effective_date > date.today()
        is_past = effective_date < date.today()

        updated_count = 0
        for fd in fees_data:
            fee_id = fd.get('id')
            new_rate = fd.get('rate')
            if fee_id and new_rate is not None:
                try:
                    fee = InspectionFee.objects.get(id=fee_id)
                    new_rate_decimal = Decimal(str(new_rate))

                    # For historical/future dates, check if a history entry already exists
                    # with the same rate to avoid spurious updates
                    existing_history = fee.history.filter(effective_date=effective_date).first()

                    needs_update = False
                    if existing_history:
                        if existing_history.rate != new_rate_decimal:
                            needs_update = True
                    else:
                        # No entry on this date - only create if rate differs from
                        # what would be active on that date
                        active_rate = fee.get_rate_for_date(effective_date)
                        if active_rate != new_rate_decimal:
                            needs_update = True

                    if needs_update:
                        FeeHistory.objects.update_or_create(
                            fee=fee, effective_date=effective_date,
                            defaults={'rate': new_rate_decimal, 'notes': notes or fd.get('notes', '')}
                        )
                        # Only update fee.rate (the "current") when effective_date is today or past
                        # AND it's the most recent history entry. Future-dated changes don't
                        # update the current rate yet.
                        if not is_future:
                            latest = fee.history.order_by('-effective_date').first()
                            if latest and latest.effective_date <= date.today():
                                fee.rate = latest.rate
                                fee.save()
                        updated_count += 1
                except InspectionFee.DoesNotExist:
                    pass
        # Clear the get_fee_rate LRU cache so updated rates take effect immediately
        try:
            from .core_views import get_fee_rate as _gfr
            _gfr.cache_clear()
        except Exception:
            pass

        return _cors(JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'effective_date': effective_date.isoformat(),
            'is_future': is_future,
            'is_past': is_past,
        }))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_react_fees_history(request):
    """Return fee change history — no auth required for Next.js."""
    from ..models import FeeHistory as _FH

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        history = _FH.objects.select_related('fee', 'created_by').order_by('-effective_date', '-id')[:200]
        data = []
        for h in history:
            data.append({
                'fee_code': h.fee.fee_code,
                'fee_name': h.fee.fee_name,
                'rate': float(h.rate),
                'effective_date': h.effective_date.isoformat() if h.effective_date else '',
                'created_by': h.created_by.get_full_name() if h.created_by else '',
                'notes': h.notes or '',
            })
        return _cors(JsonResponse({'history': data}))
    except Exception as e:
        return _cors(JsonResponse({'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Server View
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_server_view(request):
    import os
    from django.conf import settings as _s

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    try:
        from ..models import Client, FoodSafetyAgencyInspection

        # Build lookup maps for human-readable names
        client_map = {}
        for c in Client.objects.all().only('id', 'name'):
            client_map[str(c.id)] = c.name

        insp_map = {}
        insp_dates = {}  # inspection_id -> date string
        for i in FoodSafetyAgencyInspection.objects.all().only('id', 'commodity', 'product_name', 'date_of_inspection'):
            date_str = i.date_of_inspection.strftime('%d %b %Y') if i.date_of_inspection else ''
            insp_map[str(i.id)] = f"[{date_str}] {i.product_name or 'Unknown'} ({i.commodity or '?'})"
            insp_dates[str(i.id)] = date_str

        docs_root = os.path.join(_s.MEDIA_ROOT, 'docs')
        tree, total_folders, total_files = [], 0, 0
        if os.path.exists(docs_root):
            for cd in sorted(os.listdir(docs_root)):
                cp = os.path.join(docs_root, cd)
                if not os.path.isdir(cp): continue
                total_folders += 1
                client_name_raw = client_map.get(cd, f"Client {cd}")
                client_dates = set()
                cn = {'name': '', 'type': 'client', 'children': []}
                for idir in sorted(os.listdir(cp)):
                    ip = os.path.join(cp, idir)
                    if not os.path.isdir(ip): continue
                    total_folders += 1
                    insp_label = insp_map.get(idir, f"Inspection {idir}")
                    if idir in insp_dates and insp_dates[idir]:
                        client_dates.add(insp_dates[idir])
                    inode = {'name': insp_label, 'type': 'inspection', 'children': []}
                    for catd in sorted(os.listdir(ip)):
                        catp = os.path.join(ip, catd)
                        if not os.path.isdir(catp): continue
                        total_folders += 1
                        files = []
                        for f in sorted(os.listdir(catp)):
                            fp = os.path.join(catp, f)
                            if os.path.isfile(fp):
                                total_files += 1
                                rel_path = os.path.join('docs', cd, idir, catd, f)
                                files.append({'name': f, 'size': os.path.getsize(fp), 'path': rel_path})
                        inode['children'].append({'name': catd, 'type': 'category', 'files': files})
                    cn['children'].append(inode)
                # Add dates to client folder name
                if client_dates:
                    sorted_dates = sorted(client_dates)
                    if len(sorted_dates) == 1:
                        cn['name'] = f"{client_name_raw} — {sorted_dates[0]}"
                    else:
                        cn['name'] = f"{client_name_raw} — {sorted_dates[0]} to {sorted_dates[-1]}"
                else:
                    cn['name'] = client_name_raw
                tree.append(cn)
        return _cors(JsonResponse({'success': True, 'tree': tree, 'total_folders': total_folders, 'total_files': total_files}))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Settings
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_settings(request):
    import os
    from django.conf import settings as _s

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method == 'POST':
        # Trigger a manual backup
        try:
            from ..services.scheduled_backup_service import run_manual_backup as _run_backup
            success, message = _run_backup()
            return _cors(JsonResponse({'success': success, 'message': message}))
        except Exception as e:
            return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))

    try:
        from datetime import datetime as _dt
        backup_root = os.path.join(_s.BASE_DIR, 'backups')
        db_backups, excel_backups = [], []

        def _scan(directory, extensions, target_list):
            if not os.path.exists(directory):
                return
            for f in sorted(os.listdir(directory), reverse=True):
                fp = os.path.join(directory, f)
                if os.path.isfile(fp) and any(f.endswith(ext) for ext in extensions):
                    target_list.append({
                        'filename': f,
                        'size': os.path.getsize(fp),
                        'created': _dt.fromtimestamp(os.path.getmtime(fp)).isoformat(),
                    })

        # Check both root and db/ excel/ subdirectories
        _scan(backup_root, ('.sql', '.dump', '.gz', '.dump.gz', '.sqlite3'), db_backups)
        _scan(os.path.join(backup_root, 'db'), ('.sql', '.dump', '.gz', '.dump.gz', '.sqlite3'), db_backups)
        _scan(backup_root, ('.xlsx', '.csv'), excel_backups)
        _scan(os.path.join(backup_root, 'excel'), ('.xlsx', '.csv'), excel_backups)

        backup_status = 'running'

        return _cors(JsonResponse({
            'success': True,
            'database_backups': db_backups,
            'excel_exports': excel_backups,
            'backup_status': backup_status,
        }))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_download_backup(request):
    """Stream a backup file to the browser."""
    import os
    from django.conf import settings as _s
    from django.http import FileResponse, Http404

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    filename = request.GET.get('file', '')
    if not filename or '/' in filename or '\\' in filename or '..' in filename:
        raise Http404

    backup_root = os.path.join(_s.BASE_DIR, 'backups')
    search_dirs = [backup_root, os.path.join(backup_root, 'db'), os.path.join(backup_root, 'excel')]

    CONTENT_TYPES = {
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.csv': 'text/csv',
        '.sql': 'text/plain',
        '.dump': 'application/octet-stream',
        '.gz': 'application/gzip',
    }
    ext = os.path.splitext(filename)[1].lower()
    content_type = CONTENT_TYPES.get(ext, 'application/octet-stream')

    for d in search_dirs:
        candidate = os.path.join(d, filename)
        if os.path.isfile(candidate):
            response = FileResponse(open(candidate, 'rb'), content_type=content_type)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            _cors(response)
            return response
    raise Http404


# ---------------------------------------------------------------------------
#  API: Submit Ticket
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_submit_ticket(request):
    import json as _json
    from ..models import Ticket

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    if request.method != 'POST':
        return _cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))
    try:
        data = _json.loads(request.body)
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        issue_type = data.get('issue_type', '').strip()
        if not title or not description or not issue_type:
            return _cors(JsonResponse({'success': False, 'error': 'Title, description, and issue type are required'}))
        from django.contrib.auth import get_user_model
        User = get_user_model()
        ethan = User.objects.filter(username='Ethan').first()
        # Use ethan (or first admin) as created_by since API has no auth
        created_by = ethan or User.objects.filter(is_staff=True).first()
        ticket = Ticket.objects.create(title=title, description=description, issue_type=issue_type, priority=data.get('priority', 'medium'), browser_info=data.get('browser_info', ''), additional_notes=data.get('additional_notes', ''), assigned_to=ethan, created_by=created_by, status='open')
        return _cors(JsonResponse({'success': True, 'message': f'Ticket #{ticket.id} created successfully', 'ticket_id': ticket.id}))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Support Tickets Board
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_support_tickets(request):
    import json as _json
    from datetime import date as _date
    from ..models import Ticket

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        qs = Ticket.objects.select_related('created_by', 'assigned_to').order_by('-created_at')
        # Filters
        search = request.GET.get('search', '').strip()
        status_f = request.GET.get('status', '').strip()
        priority_f = request.GET.get('priority', '').strip()
        if search:
            from django.db.models import Q
            qs = qs.filter(Q(title__icontains=search) | Q(description__icontains=search))
        if status_f:
            qs = qs.filter(status=status_f)
        if priority_f:
            qs = qs.filter(priority=priority_f)

        today = _date.today()
        all_tickets = Ticket.objects.all()
        stats = {
            'open': all_tickets.filter(status='open').count(),
            'in_progress': all_tickets.filter(status='in-progress').count(),
            'resolved_today': all_tickets.filter(status='resolved', updated_at__date=today).count(),
            'high_priority': all_tickets.filter(priority='high').count(),
            'total': all_tickets.count(),
        }

        from django.contrib.auth import get_user_model
        User = get_user_model()
        all_users = list(User.objects.filter(is_active=True).values_list('username', flat=True).order_by('username'))

        tickets = []
        for t in qs:
            tickets.append({
                'id': t.id,
                'title': t.title,
                'issue_type': t.issue_type or '',
                'description': t.description or '',
                'status': t.status,
                'priority': t.priority,
                'created_by': t.created_by.get_full_name() or t.created_by.username if t.created_by else '',
                'assigned_to': t.assigned_to.username if t.assigned_to else '',
                'created_at': t.created_at.isoformat() if t.created_at else '',
                'updated_at': t.updated_at.isoformat() if t.updated_at else '',
                'browser_info': t.browser_info or '',
                'additional_notes': t.additional_notes or '',
                'affected_area': t.affected_area or '',
                'steps_to_reproduce': t.steps_to_reproduce or '',
                'expected_behavior': t.expected_behavior or '',
                'actual_behavior': t.actual_behavior or '',
                'impact_users': t.impact_users or '',
                'is_blocking': t.is_blocking or False,
            })

        return _cors(JsonResponse({'success': True, 'tickets': tickets, 'stats': stats, 'all_users': all_users}))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_support_update_status(request, ticket_id):
    import json as _json
    from ..models import Ticket

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    try:
        data = _json.loads(request.body)
        status = data.get('status', '')
        if status not in ['open', 'in-progress', 'resolved', 'closed']:
            return _cors(JsonResponse({'success': False, 'error': 'Invalid status'}, status=400))
        t = Ticket.objects.get(id=ticket_id)
        t.status = status
        t.save(update_fields=['status', 'updated_at'])
        return _cors(JsonResponse({'success': True, 'ticket_id': t.id, 'status': t.status}))
    except Ticket.DoesNotExist:
        return _cors(JsonResponse({'success': False, 'error': 'Ticket not found'}, status=404))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_support_delete_ticket(request, ticket_id):
    from ..models import Ticket

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    try:
        t = Ticket.objects.get(id=ticket_id)
        t.delete()
        return _cors(JsonResponse({'success': True}))
    except Ticket.DoesNotExist:
        return _cors(JsonResponse({'success': False, 'error': 'Ticket not found'}, status=404))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_support_create_ticket(request):
    import json as _json
    from ..models import Ticket

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    try:
        data = _json.loads(request.body)
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        if not title or not description:
            return _cors(JsonResponse({'success': False, 'error': 'Title and description required'}, status=400))
        from django.contrib.auth import get_user_model
        User = get_user_model()
        assigned_to = None
        assigned_username = data.get('assigned_to', '').strip()
        if assigned_username:
            assigned_to = User.objects.filter(username=assigned_username).first()
        ethan = User.objects.filter(username='Ethan').first()
        created_by = ethan or User.objects.filter(is_staff=True).first()
        due_date = data.get('due_date') or None
        t = Ticket.objects.create(
            title=title, description=description,
            status=data.get('status', 'open'),
            priority=data.get('priority', 'medium'),
            issue_type=data.get('issue_type', ''),
            assigned_to=assigned_to, created_by=created_by,
            due_date=due_date,
        )
        return _cors(JsonResponse({'success': True, 'ticket_id': t.id}))
    except Exception as e:
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


# ---------------------------------------------------------------------------
#  API: Debtors
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_debtors(request):
    from decimal import Decimal
    from datetime import date as _date

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))
    try:
        from ..models import XeroInvoice, XeroToken
        from .xero_views import xero_service
        xero_connected = xero_service.is_connected()
        token = XeroToken.objects.first()
        tenant_name = token.tenant_name if token else ''
        today = _date.today()
        all_invoices = XeroInvoice.objects.filter(invoice_type='ACCREC').exclude(status__in=['VOIDED', 'DELETED'])

        client_map = {}
        for inv in all_invoices:
            name = inv.contact_name or 'Unknown'
            if name not in client_map:
                client_map[name] = {'contact_name': name, 'invoices': [], 'total_invoiced': Decimal('0'), 'total_paid': Decimal('0'), 'total_outstanding': Decimal('0'), 'overdue_amount': Decimal('0'), 'aging': {'current': Decimal('0'), '1_30': Decimal('0'), '31_60': Decimal('0'), '61_90': Decimal('0'), '91_120': Decimal('0'), '120_plus': Decimal('0')}, 'invoice_count': 0, 'outstanding_count': 0, 'paid_count': 0}
            c = client_map[name]
            c['invoice_count'] += 1
            c['total_invoiced'] += inv.total
            c['invoices'].append({'invoice_number': inv.invoice_number or '', 'reference': inv.reference or '', 'date': inv.date.isoformat() if inv.date else '', 'due_date': inv.due_date.isoformat() if inv.due_date else '', 'total': float(inv.total), 'amount_paid': float(inv.amount_paid), 'amount_due': float(inv.amount_due), 'status': inv.status or '', 'days_outstanding': inv.days_outstanding})
            if inv.status == 'PAID':
                c['total_paid'] += inv.amount_paid; c['paid_count'] += 1
            elif inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'):
                c['total_outstanding'] += inv.amount_due; c['outstanding_count'] += 1
                days = inv.days_outstanding
                bucket = 'current' if days <= 0 else '1_30' if days <= 30 else '31_60' if days <= 60 else '61_90' if days <= 90 else '91_120' if days <= 120 else '120_plus'
                c['aging'][bucket] += inv.amount_due
                if inv.due_date and inv.due_date < today:
                    c['overdue_amount'] += inv.amount_due

        clients = []
        grand = {'invoiced': 0, 'paid': 0, 'outstanding': 0, 'overdue': 0}
        for c in sorted(client_map.values(), key=lambda x: float(x['total_outstanding']), reverse=True):
            grand['invoiced'] += float(c['total_invoiced']); grand['paid'] += float(c['total_paid']); grand['outstanding'] += float(c['total_outstanding']); grand['overdue'] += float(c['overdue_amount'])
            clients.append({'contact_name': c['contact_name'], 'invoice_count': c['invoice_count'], 'outstanding_count': c['outstanding_count'], 'paid_count': c['paid_count'], 'total_invoiced': float(c['total_invoiced']), 'total_paid': float(c['total_paid']), 'total_outstanding': float(c['total_outstanding']), 'overdue_amount': float(c['overdue_amount']), 'aging': {k: float(v) for k, v in c['aging'].items()}, 'invoices': c['invoices']})

        return _cors(JsonResponse({'success': True, 'clients': clients, 'grand_invoiced': grand['invoiced'], 'grand_paid': grand['paid'], 'grand_outstanding': grand['outstanding'], 'grand_overdue': grand['overdue'], 'client_count': len(clients), 'xero_connected': xero_connected, 'tenant_name': tenant_name}))
    except Exception as e:
        import traceback; traceback.print_exc()
        return _cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


from django.views.decorators.csrf import csrf_exempt as _csrf_exempt_decorator

@_csrf_exempt_decorator
def api_analytics(request):
    """Proxy for analytics dashboard data - bypasses auth for Next.js."""
    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    from .core_views import analytics_dashboard_api

    class _FakeAnalyticsUser:
        is_authenticated = True
        id = 1
        pk = 1
        username = 'api_user'
        first_name = 'Admin'
        last_name = ''
        email = 'api@localhost'
        role = 'developer'
        is_active = True
        is_staff = True
        def get_full_name(self):
            return 'Admin'
        def has_perm(self, *a, **kw):
            return True

    request.user = _FakeAnalyticsUser()
    response = analytics_dashboard_api(request)

    import json
    try:
        data = json.loads(response.content)
        from ..models import FoodSafetyAgencyInspection, InspectorTarget
        from django.db.models import Q
        from django.contrib.auth import get_user_model
        _User = get_user_model()
        _non = set(_User.objects.exclude(role='inspector').values_list('first_name', flat=True))
        for u in _User.objects.exclude(role='inspector'):
            full = f"{u.first_name} {u.last_name}".strip()
            if full: _non.add(full)
            if u.last_name: _non.add(u.last_name)
            if u.username: _non.add(u.username)
        _non.discard('')
        _non.add('admin')

        inspectors = list(FoodSafetyAgencyInspection.objects.exclude(
            Q(inspector_name__isnull=True) | Q(inspector_name='') | Q(inspector_name='Unknown') | Q(inspector_name__in=_non)
        ).values_list('inspector_name', flat=True).distinct().order_by('inspector_name'))
        commodities = list(FoodSafetyAgencyInspection.objects.exclude(
            Q(commodity__isnull=True) | Q(commodity='')
        ).values_list('commodity', flat=True).distinct().order_by('commodity'))
        years = sorted(set(
            FoodSafetyAgencyInspection.objects.exclude(date_of_inspection__isnull=True)
            .values_list('date_of_inspection__year', flat=True).distinct()
        ))

        data['filterOptions'] = {'inspectors': inspectors, 'commodities': commodities, 'years': years}

        targets = {}
        try:
            for t in InspectorTarget.objects.all():
                targets[t.inspector_name] = {
                    'eggs': t.eggs, 'poultry': t.poultry, 'raw': t.raw, 'pmp': t.pmp,
                    'raw_samples': t.raw_samples, 'pmp_samples': t.pmp_samples,
                    'total_samples': t.total_samples,
                }
        except Exception:
            pass
        data['inspectorTargets'] = targets
        data['nonInspectorNames'] = list(_non)

        # ── Salaries ──
        from ..models import InspectorSalary
        sal_data = {}
        try:
            for s in InspectorSalary.objects.all():
                sal_data[s.inspector_name.lower()] = {
                    'salary': float(s.monthly_salary),
                    'employee_number': s.employee_number or '',
                }
        except Exception:
            pass
        data['salaries'] = sal_data

        # ── Quarterly Targets ──
        from ..models import QuarterlyTarget
        qt_data = {}
        try:
            for qt in QuarterlyTarget.objects.all():
                key = f"{qt.inspector_name}_{qt.year}_Q{qt.quarter}"
                qt_data[key] = {
                    'inspector_name': qt.inspector_name,
                    'year': qt.year,
                    'quarter': qt.quarter,
                    'eggs': qt.eggs,
                    'poultry': qt.poultry,
                    'raw': qt.raw,
                    'pmp': qt.pmp,
                    'raw_samples': qt.raw_samples,
                    'pmp_samples': qt.pmp_samples,
                }
        except Exception:
            pass
        data['quarterlyTargets'] = qt_data

        from django.core.serializers.json import DjangoJSONEncoder
        response = JsonResponse(data, encoder=DjangoJSONEncoder)
    except Exception:
        pass

    return _cors(response)


def api_inspector_salaries(request):
    """Proxy for inspector salaries - bypasses auth for Next.js."""
    import json
    from ..models import InspectorSalary

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            entries = data.get('salaries', {})
            for name, salary in entries.items():
                InspectorSalary.objects.update_or_create(
                    inspector_name=name.lower(),
                    defaults={'monthly_salary': float(salary)}
                )
            # Return updated salaries
            sal_data = {}
            for s in InspectorSalary.objects.all():
                sal_data[s.inspector_name.lower()] = {
                    'salary': float(s.monthly_salary),
                    'employee_number': s.employee_number or '',
                }
            return _cors(JsonResponse({'success': True, 'salaries': sal_data}))
        except Exception as e:
            return _cors(JsonResponse({'error': str(e)}, status=500))

    # GET
    sal_data = {}
    for s in InspectorSalary.objects.all():
        sal_data[s.inspector_name.lower()] = {
            'salary': float(s.monthly_salary),
            'employee_number': s.employee_number or '',
        }
    return _cors(JsonResponse({'salaries': sal_data}))


def api_quarterly_targets(request):
    """Proxy for quarterly targets - bypasses auth for Next.js."""
    import json
    from ..models import QuarterlyTarget

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qt, created = QuarterlyTarget.objects.update_or_create(
                inspector_name=data['inspector_name'],
                year=int(data['year']),
                quarter=int(data['quarter']),
                defaults={
                    'eggs': int(data.get('eggs', 0)),
                    'poultry': int(data.get('poultry', 0)),
                    'raw': int(data.get('raw', 0)),
                    'pmp': int(data.get('pmp', 0)),
                    'raw_samples': int(data.get('raw_samples', 0)),
                    'pmp_samples': int(data.get('pmp_samples', 0)),
                }
            )
            return _cors(JsonResponse({'success': True, 'created': created}))
        except Exception as e:
            return _cors(JsonResponse({'error': str(e)}, status=500))

    # GET
    year = request.GET.get('year')
    quarter = request.GET.get('quarter')
    qs = QuarterlyTarget.objects.all()
    if year:
        qs = qs.filter(year=int(year))
    if quarter:
        qs = qs.filter(quarter=int(quarter))
    targets = {}
    for qt in qs:
        targets[qt.inspector_name] = {
            'inspector_name': qt.inspector_name,
            'year': qt.year,
            'quarter': qt.quarter,
            'eggs': qt.eggs,
            'poultry': qt.poultry,
            'raw': qt.raw,
            'pmp': qt.pmp,
            'raw_samples': qt.raw_samples,
            'pmp_samples': qt.pmp_samples,
        }
    return _cors(JsonResponse({'targets': targets}))


# ---------------------------------------------------------------------------
#  API: Get / Edit Inspection Group (for Next.js edit page)
# ---------------------------------------------------------------------------

def _insp_cors(r):
    r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
    r['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    r['Access-Control-Allow-Headers'] = 'Content-Type'
    return r


@_csrf_exempt
def api_get_inspection_group(request, pk):
    """Return the full data for an inspection group, keyed by the first product pk."""
    from ..models import FoodSafetyAgencyInspection as _Insp, InspectionGroup as _Group

    if request.method == 'OPTIONS':
        return _insp_cors(JsonResponse({'ok': True}))

    try:
        insp = _Insp.objects.select_related('inspection_group', 'client').get(pk=pk)
        group = insp.inspection_group

        if group:
            all_inspections = list(
                _Insp.objects.filter(inspection_group=group).order_by('commodity', 'id')
            )
            km = float(group.km_traveled) if group.km_traveled else 0
            hours = float(group.hours) if group.hours else 0
            client_name = group.client_name or ''
            town = group.town or ''
            inspector_name = group.inspector_name or ''
            date_of_inspection = group.date_of_inspection.isoformat() if group.date_of_inspection else ''
            corporate_group = group.corporate_group or ''
            group_type = group.group_type or ''
            facility_type = group.facility_type or ''
            additional_email = group.additional_email or ''
            comment = group.comment or ''
            travel_start_time = group.travel_start_time.strftime('%H:%M') if group.travel_start_time else ''
            travel_end_time = group.travel_end_time.strftime('%H:%M') if group.travel_end_time else ''
        else:
            all_inspections = [insp]
            km = float(insp.km_traveled) if insp.km_traveled else 0
            hours = float(insp.hours) if insp.hours else 0
            client_name = insp.client_name or ''
            town = insp.town or ''
            inspector_name = insp.inspector_name or ''
            date_of_inspection = insp.date_of_inspection.isoformat() if insp.date_of_inspection else ''
            corporate_group = insp.corporate_group or ''
            group_type = insp.group_type or ''
            facility_type = insp.facility_type or ''
            additional_email = insp.additional_email or ''
            comment = insp.comment or ''
            travel_start_time = ''
            travel_end_time = ''

        products = []
        for p in all_inspections:
            products.append({
                'id': p.id,
                'commodity': p.commodity or '',
                'product_name': p.product_name or '',
                'product_class': p.product_class or '',
                'lab': p.lab or '',
                'is_sample_taken': bool(p.is_sample_taken),
                'fat': bool(p.fat),
                'protein': bool(p.protein),
                'calcium': bool(p.calcium),
                'dna': bool(p.dna),
                'needs_retest': p.needs_retest or 'NO',
                'bought_sample': float(p.bought_sample) if p.bought_sample else 0,
            })

        # Get primary client email
        client_email = ''
        if insp.client:
            client_email = insp.client.email or insp.client.manual_email or ''

        return _insp_cors(JsonResponse({
            'success': True,
            'inspection_id': pk,
            'group_id': group.id if group else None,
            'client_name': client_name,
            'town': town,
            'inspector_name': inspector_name,
            'date_of_inspection': date_of_inspection,
            'corporate_group': corporate_group,
            'group_type': group_type,
            'facility_type': facility_type,
            'client_email': client_email,
            'additional_email': additional_email,
            'comment': comment,
            'km_traveled': km,
            'hours': hours,
            'travel_start_time': travel_start_time,
            'travel_end_time': travel_end_time,
            'products': products,
            'is_occurrence_report': bool(insp.is_occurrence_report),
        }))
    except _Insp.DoesNotExist:
        return _insp_cors(JsonResponse({'success': False, 'error': 'Inspection not found'}, status=404))
    except Exception as e:
        return _insp_cors(JsonResponse({'success': False, 'error': str(e)}, status=500))


@_csrf_exempt
def api_edit_inspection_group(request):
    """JSON POST endpoint to update an inspection group from the Next.js edit page."""
    from ..models import FoodSafetyAgencyInspection as _Insp, InspectionGroup as _Group
    from collections import defaultdict
    import json as _json
    from django.db import transaction
    from django.db.models import Min
    from datetime import date as _date

    if request.method == 'OPTIONS':
        return _insp_cors(JsonResponse({'ok': True}))

    if request.method != 'POST':
        return _insp_cors(JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405))

    try:
        data = _json.loads(request.body)
        inspection_id = data.get('inspection_id')
        if not inspection_id:
            return _insp_cors(JsonResponse({'success': False, 'error': 'inspection_id is required'}))

        # Quick update for approved_status only
        approved_only = data.get('approved_status')
        if approved_only and len(data) <= 2:  # Only inspection_id + approved_status
            try:
                group = _Group.objects.get(pk=inspection_id)
            except _Group.DoesNotExist:
                return _insp_cors(JsonResponse({'success': False, 'error': 'Group not found'}))
            _Insp.objects.filter(inspection_group=group).update(approved_status=approved_only)
            return _insp_cors(JsonResponse({'success': True}))

        # Try as InspectionGroup first, then as FoodSafetyAgencyInspection
        try:
            group = _Group.objects.get(pk=inspection_id)
            insp = _Insp.objects.filter(inspection_group=group).select_related('client').first()
        except _Group.DoesNotExist:
            insp = _Insp.objects.select_related('inspection_group', 'client').get(pk=inspection_id)
            group = insp.inspection_group

        # Parse date
        from datetime import datetime as _datetime
        date_str = data.get('date_of_inspection', '')
        try:
            date_obj = _datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else insp.date_of_inspection
        except ValueError:
            date_obj = insp.date_of_inspection

        # Common fields from POST
        client_name = data.get('client_name', insp.client_name or '').strip()
        town = data.get('town', insp.town or '').strip()
        inspector_name = data.get('inspector_name', insp.inspector_name or '').strip()
        corporate_group = data.get('corporate_group', '').strip()
        group_type = data.get('group_type', '').strip()
        facility_type = data.get('facility_type', '').strip()
        additional_email = data.get('additional_email', '').strip()
        client_email = data.get('client_email', '').strip()
        comment = data.get('comment', '').strip()
        km_traveled = float(data.get('km_traveled', 0) or 0)
        hours = float(data.get('hours', 0) or 0)
        travel_start_time = data.get('travel_start_time', '') or ''
        travel_end_time = data.get('travel_end_time', '') or ''
        products_data = data.get('products', [])

        with transaction.atomic():
            # Update parent InspectionGroup
            if group:
                group.client_name = client_name
                group.town = town
                group.inspector_name = inspector_name
                group.date_of_inspection = date_obj
                group.corporate_group = corporate_group
                group.group_type = group_type
                group.facility_type = facility_type
                group.additional_email = additional_email
                group.comment = comment
                group.km_traveled = km_traveled
                group.hours = hours
                if travel_start_time:
                    try:
                        from datetime import time as _time
                        parts = travel_start_time.split(':')
                        group.travel_start_time = _time(int(parts[0]), int(parts[1]))
                    except Exception:
                        pass
                if travel_end_time:
                    try:
                        from datetime import time as _time
                        parts = travel_end_time.split(':')
                        group.travel_end_time = _time(int(parts[0]), int(parts[1]))
                    except Exception:
                        pass
                group.save()

                # Update client email if provided
                if group.client and additional_email:
                    # First email in the list is the primary
                    all_emails = [e.strip() for e in additional_email.replace(',', ';').split(';') if e.strip()]
                    if all_emails:
                        group.client.email = all_emails[0]
                        group.client.save(update_fields=['email'])

                related_inspections = list(_Insp.objects.filter(inspection_group=group).order_by('commodity', 'id'))
            else:
                related_inspections = [insp]

            if products_data:
                # Group products_data by commodity
                products_by_commodity = defaultdict(list)
                for p in products_data:
                    if p.get('commodity'):
                        products_by_commodity[p['commodity']].append(p)

                # Group existing inspections by commodity
                existing_by_commodity = defaultdict(list)
                for rel in related_inspections:
                    existing_by_commodity[rel.commodity].append(rel)

                to_update = []
                to_delete = []
                to_create = []

                # Match products → existing by commodity index
                for commodity, prods in products_by_commodity.items():
                    existing = existing_by_commodity.get(commodity, [])
                    for idx, prod in enumerate(prods):
                        if idx < len(existing):
                            rel = existing[idx]
                            rel.product_name = prod.get('product_name', '')
                            rel.product_class = prod.get('product_class', '')
                            rel.lab = prod.get('lab', '')
                            rel.is_sample_taken = prod.get('is_sample_taken', False)
                            rel.fat = prod.get('fat', False)
                            rel.protein = prod.get('protein', False)
                            rel.calcium = prod.get('calcium', False)
                            rel.dna = prod.get('dna', False)
                            rel.needs_retest = prod.get('needs_retest', 'NO')
                            rel.bought_sample = prod.get('bought_sample', 0)
                            rel.km_traveled = km_traveled
                            rel.hours = hours
                            rel.client_name = client_name
                            rel.town = town
                            rel.inspector_name = inspector_name
                            rel.date_of_inspection = date_obj
                            rel.corporate_group = corporate_group
                            rel.group_type = group_type
                            rel.facility_type = facility_type
                            rel.additional_email = additional_email
                            to_update.append(rel)
                        else:
                            prod['_commodity'] = commodity
                            to_create.append(prod)
                    # Do NOT delete extra existing products within same commodity
                    # They may be legitimate products not sent by the edit form

                # Only delete products if explicitly flagged for removal
                # Do NOT delete commodities just because they weren't sent in the payload
                # This prevents accidental deletion when the edit form doesn't send all products
                explicitly_removed = data.get('removed_commodities', [])
                if explicitly_removed:
                    for commodity, existing in existing_by_commodity.items():
                        if commodity in explicitly_removed:
                            to_delete.extend(existing)

                for rel in to_delete:
                    rel.delete()

                for rel in to_update:
                    rel.save()

                # Create new products
                for prod in to_create:
                    min_remote = _Insp.objects.filter(is_manual=True).aggregate(Min('remote_id'))['remote_id__min']
                    new_remote_id = -1 if (min_remote is None or min_remote >= 0) else min_remote - 1
                    _Insp.objects.create(
                        inspection_group=group,
                        remote_id=new_remote_id,
                        client=insp.client,
                        client_name=client_name,
                        date_of_inspection=date_obj,
                        commodity=prod['_commodity'],
                        product_name=prod.get('product_name', ''),
                        product_class=prod.get('product_class', ''),
                        lab=prod.get('lab', ''),
                        is_sample_taken=prod.get('is_sample_taken', False),
                        fat=prod.get('fat', False),
                        protein=prod.get('protein', False),
                        calcium=prod.get('calcium', False),
                        dna=prod.get('dna', False),
                        needs_retest=prod.get('needs_retest', 'NO'),
                        bought_sample=prod.get('bought_sample', 0),
                        km_traveled=km_traveled,
                        hours=hours,
                        inspector_name=inspector_name,
                        town=town,
                        corporate_group=corporate_group,
                        group_type=group_type,
                        facility_type=facility_type,
                        additional_email=additional_email,
                        is_manual=True,
                    )
            else:
                # No products data — update common fields on all existing inspections
                for rel in related_inspections:
                    rel.client_name = client_name
                    rel.town = town
                    rel.inspector_name = inspector_name
                    rel.date_of_inspection = date_obj
                    rel.corporate_group = corporate_group
                    rel.group_type = group_type
                    rel.facility_type = facility_type
                    rel.additional_email = additional_email
                    rel.km_traveled = km_traveled
                    rel.hours = hours
                    rel.save()

            # Re-sequence
            if group:
                for seq_num, rel in enumerate(
                    _Insp.objects.filter(inspection_group=group).order_by('id'), start=1
                ):
                    rel.inspection_sequence = seq_num
                    rel.save(update_fields=['inspection_sequence'])

            # Clear cache
            from django.core.cache import cache as _fc
            _fc.delete('add_inspection_form_data')
            _fc.delete('page_clients_status_cache')

        return _insp_cors(JsonResponse({'success': True, 'message': f'Inspection for {client_name} updated successfully'}))

    except _Insp.DoesNotExist:
        return _insp_cors(JsonResponse({'success': False, 'error': 'Inspection not found'}, status=404))
    except Exception as e:
        import traceback
        return _insp_cors(JsonResponse({'success': False, 'error': str(e), 'trace': traceback.format_exc()}, status=500))


@_csrf_exempt
def api_admin_analytics(request):
    """Admin analytics dashboard — overview stats for the admin role."""
    from ..models import FoodSafetyAgencyInspection as _I, Client as _C, InspectionGroup as _G
    from django.db.models import Count, Q
    import datetime
    from django.utils import timezone as _tz

    def _cors(r):
        r['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        r['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        r['Access-Control-Allow-Headers'] = 'Content-Type'
        return r

    if request.method == 'OPTIONS':
        return _cors(JsonResponse({'ok': True}))

    try:
        today = _tz.now().date()

        # ── Filters ──────────────────────────────────────────────────────────
        filter_inspector = request.GET.get('inspector', '').strip()
        filter_date_from = request.GET.get('date_from', '').strip()
        filter_date_to = request.GET.get('date_to', '').strip()

        groups_qs = _G.objects.all()
        inspections_qs = _I.objects.all()
        if filter_inspector:
            groups_qs = groups_qs.filter(inspector_name__iexact=filter_inspector)
            inspections_qs = inspections_qs.filter(inspector_name__iexact=filter_inspector)
        if filter_date_from:
            groups_qs = groups_qs.filter(date_of_inspection__gte=filter_date_from)
            inspections_qs = inspections_qs.filter(date_of_inspection__gte=filter_date_from)
        if filter_date_to:
            groups_qs = groups_qs.filter(date_of_inspection__lte=filter_date_to)
            inspections_qs = inspections_qs.filter(date_of_inspection__lte=filter_date_to)

        # Get all inspector names for the filter dropdown
        all_inspectors = list(
            _G.objects.exclude(inspector_name__isnull=True).exclude(inspector_name='')
            .values_list('inspector_name', flat=True).distinct().order_by('inspector_name')
        )

        # ── Summary counts ────────────────────────────────────────────────────
        total_clients     = _C.objects.count()
        total_groups      = groups_qs.count()
        total_inspections = inspections_qs.count()

        # This month
        first_of_month = today.replace(day=1)
        this_month_groups = groups_qs.filter(date_of_inspection__gte=first_of_month).count()

        # Sent / approved counts
        total_sent     = groups_qs.filter(inspections__sent_date__isnull=False).distinct().count()
        total_approved = groups_qs.filter(inspections__approved_status='APPROVED').distinct().count()

        # ── Compliance ────────────────────────────────────────────────────────
        non_compliant_groups = groups_qs.filter(
            inspections__is_direction_present_for_this_inspection=True
        ).distinct().count()
        compliance_rate = round(
            (1 - non_compliant_groups / total_groups) * 100, 1
        ) if total_groups else 0.0

        # ── Monthly trend — last 6 months ────────────────────────────────────
        monthly = []
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            first = datetime.date(y, m, 1)
            last = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1) if m < 12 else datetime.date(y + 1, 1, 1) - datetime.timedelta(days=1)
            count = groups_qs.filter(date_of_inspection__gte=first, date_of_inspection__lte=last).count()
            monthly.append({'month': first.strftime('%b %Y'), 'count': count})

        # ── Top inspectors ────────────────────────────────────────────────────
        top_inspectors = list(
            groups_qs.exclude(inspector_name='').exclude(inspector_name__isnull=True)
            .values('inspector_name')
            .annotate(count=Count('id'))
            .order_by('-count')[:8]
        )

        # ── Group type breakdown ─────────────────────────────────────────────
        group_type_data = list(
            groups_qs.exclude(group_type='').exclude(group_type__isnull=True)
            .values('group_type').annotate(count=Count('id')).order_by('-count')[:10]
        )

        # ── Corporate group breakdown ────────────────────────────────────────
        corporate_data = list(
            groups_qs.exclude(corporate_group='').exclude(corporate_group__isnull=True)
            .values('corporate_group').annotate(count=Count('id')).order_by('-count')[:10]
        )

        # ── Sent rate by month ───────────────────────────────────────────────
        sent_monthly = []
        for entry in monthly:
            dt = datetime.datetime.strptime(entry['month'], '%b %Y')
            first = datetime.date(dt.year, dt.month, 1)
            last = datetime.date(dt.year, dt.month + 1, 1) - datetime.timedelta(days=1) if dt.month < 12 else datetime.date(dt.year + 1, 1, 1) - datetime.timedelta(days=1)
            sent = groups_qs.filter(date_of_inspection__gte=first, date_of_inspection__lte=last, inspections__sent_date__isnull=False).distinct().count()
            sent_monthly.append({'month': entry['month'], 'sent': sent, 'total': entry['count']})

        return _cors(JsonResponse({
            'success': True,
            'total_clients':      total_clients,
            'total_groups':       total_groups,
            'total_inspections':  total_inspections,
            'this_month_groups':  this_month_groups,
            'total_sent':         total_sent,
            'total_approved':     total_approved,
            'compliance_rate':    compliance_rate,
            'non_compliant':      non_compliant_groups,
            'monthly':            monthly,
            'sent_monthly':       sent_monthly,
            'top_inspectors':     top_inspectors,
            'group_type_data':    group_type_data,
            'corporate_data':     corporate_data,
            'all_inspectors':     all_inspectors,
        }))

    except Exception as e:
        import traceback
        return _cors(JsonResponse({'success': False, 'error': str(e), 'trace': traceback.format_exc()}, status=500))


# ---------------------------------------------------------------------------
#  Inspector Mappings API  (JSON, for React frontend)
# ---------------------------------------------------------------------------
@_csrf_exempt
def api_inspector_mappings(request):
    """JSON API endpoint for inspector mappings CRUD — used by Next.js frontend."""
    from ..models import InspectorMapping

    def _cors_json(data, status=200):
        resp = JsonResponse(data, status=status, safe=False)
        resp['Access-Control-Allow-Origin'] = 'http://localhost:3000'
        resp['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        resp['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp

    if request.method == 'OPTIONS':
        return _cors_json({'ok': True})

    # ---------- GET: list all mappings ----------
    if request.method == 'GET':
        try:
            mappings = InspectorMapping.objects.all().order_by('inspector_name')
            data = [
                {
                    'id': m.pk,
                    'inspector_id': m.inspector_id,
                    'inspector_name': m.inspector_name,
                    'is_active': m.is_active,
                    'created_at': m.created_at.isoformat() if m.created_at else None,
                    'updated_at': m.updated_at.isoformat() if m.updated_at else None,
                }
                for m in mappings
            ]
            return _cors_json({'success': True, 'mappings': data})
        except Exception as e:
            return _cors_json({'success': False, 'error': str(e)}, status=500)

    # ---------- POST: add / edit / delete ----------
    if request.method == 'POST':
        import json as _json
        try:
            body = _json.loads(request.body)
        except Exception:
            return _cors_json({'success': False, 'error': 'Invalid JSON body'}, status=400)

        action = body.get('action', '')

        # ---- ADD ----
        if action == 'add':
            inspector_id = body.get('inspector_id')
            inspector_name = (body.get('inspector_name') or '').strip()
            is_active = body.get('is_active', True)

            if not inspector_id or not inspector_name:
                return _cors_json({'success': False, 'error': 'Inspector ID and name are required.'}, status=400)

            if InspectorMapping.objects.filter(inspector_id=inspector_id).exists():
                return _cors_json({'success': False, 'error': f'Inspector ID {inspector_id} already exists.'}, status=400)

            try:
                m = InspectorMapping.objects.create(
                    inspector_id=int(inspector_id),
                    inspector_name=inspector_name,
                    is_active=bool(is_active),
                )
                return _cors_json({'success': True, 'message': f'Inspector mapping "{inspector_name}" added.', 'id': m.pk})
            except Exception as e:
                return _cors_json({'success': False, 'error': str(e)}, status=500)

        # ---- EDIT ----
        if action == 'edit':
            pk = body.get('pk') or body.get('id')
            if not pk:
                return _cors_json({'success': False, 'error': 'Missing mapping ID.'}, status=400)

            try:
                m = InspectorMapping.objects.get(pk=int(pk))
            except InspectorMapping.DoesNotExist:
                return _cors_json({'success': False, 'error': 'Mapping not found.'}, status=404)

            inspector_id = body.get('inspector_id')
            inspector_name = (body.get('inspector_name') or '').strip()
            is_active = body.get('is_active', m.is_active)

            if inspector_id is not None:
                # Check uniqueness if changing the inspector_id
                if int(inspector_id) != m.inspector_id and InspectorMapping.objects.filter(inspector_id=int(inspector_id)).exists():
                    return _cors_json({'success': False, 'error': f'Inspector ID {inspector_id} already exists.'}, status=400)
                m.inspector_id = int(inspector_id)

            if inspector_name:
                m.inspector_name = inspector_name

            m.is_active = bool(is_active)
            m.save()
            return _cors_json({'success': True, 'message': f'Inspector mapping "{m.inspector_name}" updated.'})

        # ---- DELETE ----
        if action == 'delete':
            pk = body.get('pk') or body.get('id')
            if not pk:
                return _cors_json({'success': False, 'error': 'Missing mapping ID.'}, status=400)

            try:
                m = InspectorMapping.objects.get(pk=int(pk))
                name = m.inspector_name
                m.delete()
                return _cors_json({'success': True, 'message': f'Inspector mapping "{name}" deleted.'})
            except InspectorMapping.DoesNotExist:
                return _cors_json({'success': False, 'error': 'Mapping not found.'}, status=404)

        return _cors_json({'success': False, 'error': f'Unknown action: {action}'}, status=400)

    return _cors_json({'success': False, 'error': 'Method not allowed.'}, status=405)


@_csrf_exempt
def api_onedrive_view(request):
    """JSON API for OneDrive folder listing — used by Next.js frontend."""
    import os, json as _json
    from datetime import datetime
    from django.conf import settings as _settings

    data = {'connected': False, 'folders': [], 'total_files': 0, 'error': None, 'token_refreshed': False}

    try:
        token_file = os.path.join(_settings.BASE_DIR, 'onedrive_tokens.json')
        if os.path.exists(token_file):
            with open(token_file, 'r') as f:
                tokens = _json.load(f)

            access_token = tokens.get('access_token')
            expires_at = tokens.get('expires_at', 0)
            current_time = datetime.now().timestamp()

            if access_token and current_time < expires_at:
                data['connected'] = True
            elif access_token:
                data['token_refreshed'] = True
                data['connected'] = True
            else:
                data['error'] = 'No valid access token found'
        else:
            data['error'] = 'OneDrive not connected. Please connect in Settings.'

        if data['connected']:
            from ..services.onedrive_direct_service import OneDriveDirectUploadService
            svc = OneDriveDirectUploadService()
            if svc.authenticate_onedrive():
                folders = svc.list_folders_in_onedrive(None)
                if folders:
                    data['folders'] = folders
                    data['total_files'] = sum(f.get('file_count', 0) for f in folders)
                else:
                    data['error'] = 'No folders found in OneDrive'
            else:
                data['error'] = 'Failed to authenticate with OneDrive.'
                data['connected'] = False
    except Exception as e:
        data['error'] = f'Error accessing OneDrive: {str(e)}'
        data['connected'] = False

    return json_response(data)


@_csrf_exempt
def api_convert_docx_to_pdf(request):
    """Convert an uploaded .docx file to PDF using docx2pdf (Windows/Word) or LibreOffice."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    import tempfile, os, subprocess
    tmp_dir = tempfile.mkdtemp()
    docx_path = os.path.join(tmp_dir, 'report.docx')
    pdf_path = os.path.join(tmp_dir, 'report.pdf')

    try:
        with open(docx_path, 'wb') as f:
            for chunk in uploaded.chunks():
                f.write(chunk)

        # Try docx2pdf first (uses MS Word on Windows)
        try:
            from docx2pdf import convert
            convert(docx_path, pdf_path)
        except Exception:
            # Fallback: LibreOffice headless
            try:
                subprocess.run([
                    'soffice', '--headless', '--convert-to', 'pdf',
                    '--outdir', tmp_dir, docx_path
                ], check=True, timeout=30)
            except Exception as e2:
                return JsonResponse({'error': f'PDF conversion failed: {str(e2)}'}, status=500)

        if not os.path.exists(pdf_path):
            return JsonResponse({'error': 'PDF conversion produced no output'}, status=500)

        with open(pdf_path, 'rb') as f:
            pdf_bytes = f.read()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Analytics_Report.pdf"'
        return response
    finally:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)


@_csrf_exempt
def api_send_occurrence_email(request):
    """Send a notification email when an occurrence report is submitted."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    from datetime import datetime as dt_parser
    from django.core.mail import EmailMessage

    try:
        # Handle both JSON and FormData
        if request.content_type and 'multipart' in request.content_type:
            client_name = request.POST.get('client_name', '').strip()
            town = request.POST.get('town', '').strip()
            inspection_date = request.POST.get('date_of_inspection', '').strip()
            inspector_name = request.POST.get('inspector_name', '').strip()
            description = request.POST.get('description', '').strip()
            corporate_group = request.POST.get('corporate_group', '').strip()
            uploaded_file = request.FILES.get('file')
        else:
            import json
            data = json.loads(request.body)
            client_name = data.get('client_name', '').strip()
            town = data.get('town', '').strip()
            inspection_date = data.get('date_of_inspection', '').strip()
            inspector_name = data.get('inspector_name', '').strip()
            description = data.get('description', '').strip()
            corporate_group = data.get('corporate_group', '').strip()
            uploaded_file = None

        try:
            formatted_date = dt_parser.strptime(inspection_date, '%Y-%m-%d').strftime('%d %B %Y')
        except (ValueError, TypeError):
            formatted_date = inspection_date

        subject = f'Occurrence Report – {client_name} – {formatted_date}'
        html_message = f"""
<div style="font-family: Calibri, Arial, sans-serif; font-size: 14px; color: #333; line-height: 1.6;">
    <p>Good day,</p>

    <p>An occurrence report has been submitted and requires your attention for potential legal claims processing.</p>

    <table style="border-collapse: collapse; width: 100%; max-width: 600px; margin: 16px 0;">
        <tr style="background: #f8f9fa;">
            <td style="padding: 10px 14px; font-weight: 600; border: 1px solid #e5e7eb; width: 160px;">Client</td>
            <td style="padding: 10px 14px; border: 1px solid #e5e7eb;">{client_name}</td>
        </tr>
        <tr>
            <td style="padding: 10px 14px; font-weight: 600; border: 1px solid #e5e7eb;">Town</td>
            <td style="padding: 10px 14px; border: 1px solid #e5e7eb;">{town}</td>
        </tr>
        <tr style="background: #f8f9fa;">
            <td style="padding: 10px 14px; font-weight: 600; border: 1px solid #e5e7eb;">Corporate Group</td>
            <td style="padding: 10px 14px; border: 1px solid #e5e7eb;">{corporate_group or 'N/A'}</td>
        </tr>
        <tr>
            <td style="padding: 10px 14px; font-weight: 600; border: 1px solid #e5e7eb;">Date of Visit</td>
            <td style="padding: 10px 14px; border: 1px solid #e5e7eb;">{formatted_date}</td>
        </tr>
        <tr style="background: #f8f9fa;">
            <td style="padding: 10px 14px; font-weight: 600; border: 1px solid #e5e7eb;">Inspector</td>
            <td style="padding: 10px 14px; border: 1px solid #e5e7eb;">{inspector_name}</td>
        </tr>
    </table>

    <p><strong>Occurrence Description:</strong></p>
    <div style="background: #fff8f0; border-left: 4px solid #f59e0b; padding: 14px 18px; margin: 12px 0 20px; border-radius: 0 8px 8px 0; white-space: pre-wrap;">
        {description if description else '<em style="color: #9ca3af;">No description provided.</em>'}
    </div>

    <p>Please review this occurrence report at your earliest convenience. This record has been logged in the APS system and may be required for legal claims or compliance follow-up.</p>

    <p>Kind Regards / Vriendelike Groete<br>
    <strong>APS Inspection System</strong></p>
</div>"""

        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=['nicole.bergh@afsq.co.za'],
            reply_to=[settings.DEFAULT_FROM_EMAIL],
        )
        email.content_subtype = 'html'

        if uploaded_file:
            email.attach(uploaded_file.name, uploaded_file.read(), uploaded_file.content_type or 'application/pdf')

        email.send()

        return JsonResponse({'success': True, 'message': 'Occurrence notification sent'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
