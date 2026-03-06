"""
Xero Accounting Integration Views
OAuth2 connect/callback, invoice sync, and status endpoints.
"""
import logging
import io
from datetime import timedelta, date
from decimal import Decimal
from django.shortcuts import redirect, render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q, Count

from ..services import xero_service
from ..models import XeroToken, XeroInvoice

logger = logging.getLogger(__name__)


@login_required
def xero_connect(request):
    """Redirect user to Xero OAuth2 authorization page."""
    if request.user.role not in ('admin', 'super_admin', 'developer'):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    url = xero_service.get_authorization_url()
    return redirect(url)


@login_required
def xero_callback(request):
    """Handle Xero OAuth2 callback after user authorizes."""
    code = request.GET.get('code')
    error = request.GET.get('error')

    if error:
        logger.error(f'Xero OAuth error: {error}')
        return redirect('/analytics-dashboard/#financial')

    if not code:
        logger.error('Xero callback: no code received')
        return redirect('/analytics-dashboard/#financial')

    try:
        # Exchange code for tokens
        token_data = xero_service.exchange_code_for_tokens(code)

        access_token = token_data['access_token']
        refresh_token = token_data.get('refresh_token', '')
        expires_in = token_data.get('expires_in', 1800)

        # Get tenant info
        connections = xero_service.get_tenant_connections(access_token)
        tenant_id = ''
        tenant_name = ''
        if connections:
            tenant_id = connections[0].get('tenantId', '')
            tenant_name = connections[0].get('tenantName', '')

        # Store token (replace any existing)
        XeroToken.objects.all().delete()
        XeroToken.objects.create(
            access_token=access_token,
            refresh_token=refresh_token,
            expires_at=timezone.now() + timedelta(seconds=expires_in),
            tenant_id=tenant_id,
            tenant_name=tenant_name,
        )

        logger.info(f'Xero connected successfully to: {tenant_name}')

    except Exception as e:
        logger.error(f'Xero callback error: {e}')

    return redirect('/analytics-dashboard/#financial')


@login_required
def xero_disconnect(request):
    """Disconnect from Xero by removing stored tokens."""
    if request.user.role not in ('admin', 'super_admin', 'developer'):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    XeroToken.objects.all().delete()
    logger.info('Xero disconnected')
    return JsonResponse({'status': 'disconnected'})


@login_required
def xero_status(request):
    """Check Xero connection status."""
    token = XeroToken.objects.first()
    if not token:
        return JsonResponse({'connected': False})

    connected = xero_service.is_connected()
    return JsonResponse({
        'connected': connected,
        'tenant_name': token.tenant_name if connected else '',
        'expires_at': token.expires_at.isoformat() if connected else '',
    })


@login_required
@require_POST
def xero_sync_invoices(request):
    """Trigger a sync of invoices from Xero."""
    if request.user.role not in ('admin', 'super_admin', 'developer'):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        result = xero_service.sync_invoices_to_db()
        return JsonResponse({
            'status': 'success',
            'synced': result['synced'],
            'errors': result['errors'],
            'total': result['total'],
        })
    except Exception as e:
        logger.error(f'Xero sync error: {e}')
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def xero_invoices_list(request):
    """Return synced Xero invoices as JSON for the dashboard."""
    invoices = XeroInvoice.objects.all()

    # Optional filters
    status = request.GET.get('status')
    if status:
        invoices = invoices.filter(status=status)

    data = []
    aging_summary = {'Current': 0, '1-30': 0, '31-60': 0, '61-90': 0, '91-120': 0, '120+': 0}

    for inv in invoices:
        bucket = inv.aging_bucket
        if inv.status != 'PAID':
            aging_summary[bucket] = aging_summary.get(bucket, 0) + float(inv.amount_due)

        data.append({
            'invoice_number': inv.invoice_number,
            'contact_name': inv.contact_name,
            'reference': inv.reference,
            'status': inv.status,
            'total': float(inv.total),
            'amount_due': float(inv.amount_due),
            'amount_paid': float(inv.amount_paid),
            'date': inv.date.isoformat() if inv.date else '',
            'due_date': inv.due_date.isoformat() if inv.due_date else '',
            'days_outstanding': inv.days_outstanding,
            'aging_bucket': bucket,
        })

    return JsonResponse({
        'invoices': data,
        'aging_summary': aging_summary,
        'total_outstanding': sum(aging_summary.values()),
        'count': len(data),
    })


@login_required
def debtors_page(request):
    """Debtors page - shows clients with paid/outstanding/overdue from Xero invoices."""
    if request.user.role not in ('admin', 'super_admin', 'developer'):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Permission denied')

    # Check Xero connection
    xero_connected = xero_service.is_connected()
    token = XeroToken.objects.first()
    tenant_name = token.tenant_name if token else ''

    today = date.today()

    # Get all ACCREC (sales) invoices - exclude VOIDED and DELETED
    all_invoices = XeroInvoice.objects.filter(
        invoice_type='ACCREC'
    ).exclude(status__in=['VOIDED', 'DELETED'])

    # Build client-level data: group invoices by contact_name
    client_map = {}
    for inv in all_invoices:
        name = inv.contact_name or 'Unknown'
        if name not in client_map:
            client_map[name] = {
                'contact_name': name,
                'contact_id': inv.contact_id,
                'invoices': [],
                'total_invoiced': Decimal('0'),
                'total_paid': Decimal('0'),
                'total_outstanding': Decimal('0'),
                'overdue_amount': Decimal('0'),
                'oldest_overdue_days': 0,
                'aging': {
                    'current': Decimal('0'),
                    '1_30': Decimal('0'),
                    '31_60': Decimal('0'),
                    '61_90': Decimal('0'),
                    '91_120': Decimal('0'),
                    '120_plus': Decimal('0'),
                },
                'invoice_count': 0,
                'outstanding_count': 0,
                'paid_count': 0,
            }

        c = client_map[name]
        c['invoices'].append(inv)
        c['invoice_count'] += 1
        c['total_invoiced'] += inv.total

        if inv.status == 'PAID':
            c['total_paid'] += inv.amount_paid
            c['paid_count'] += 1
        elif inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'):
            c['total_outstanding'] += inv.amount_due
            c['outstanding_count'] += 1

            # Aging bucket
            days = inv.days_outstanding
            if days <= 0:
                c['aging']['current'] += inv.amount_due
            elif days <= 30:
                c['aging']['1_30'] += inv.amount_due
            elif days <= 60:
                c['aging']['31_60'] += inv.amount_due
            elif days <= 90:
                c['aging']['61_90'] += inv.amount_due
            elif days <= 120:
                c['aging']['91_120'] += inv.amount_due
            else:
                c['aging']['120_plus'] += inv.amount_due

            # Overdue tracking
            if inv.due_date and inv.due_date < today:
                c['overdue_amount'] += inv.amount_due
                c['oldest_overdue_days'] = max(c['oldest_overdue_days'], days)

    # Determine worst aging bucket per client
    for name, c in client_map.items():
        if c['aging']['120_plus'] > 0:
            c['worst_bucket'] = '120+'
        elif c['aging']['91_120'] > 0:
            c['worst_bucket'] = '91-120'
        elif c['aging']['61_90'] > 0:
            c['worst_bucket'] = '61-90'
        elif c['aging']['31_60'] > 0:
            c['worst_bucket'] = '31-60'
        elif c['aging']['1_30'] > 0:
            c['worst_bucket'] = '1-30'
        else:
            c['worst_bucket'] = 'Current'

        # Sort invoices by due_date
        c['invoices'].sort(key=lambda x: x.due_date or date.min)

    # Sort clients by outstanding amount descending
    clients = sorted(client_map.values(), key=lambda x: x['total_outstanding'], reverse=True)

    # Apply filters
    filter_bucket = request.GET.get('aging', '')
    filter_status = request.GET.get('status', '')
    search_q = request.GET.get('q', '').strip()
    filter_inv_status = request.GET.get('inv_status', '')
    filter_date_from = request.GET.get('date_from', '')
    filter_date_to = request.GET.get('date_to', '')
    filter_due_from = request.GET.get('due_from', '')
    filter_due_to = request.GET.get('due_to', '')
    sort_by = request.GET.get('sort', '')

    # Parse date filters
    date_from = None
    date_to = None
    due_from = None
    due_to = None
    try:
        if filter_date_from:
            date_from = date.fromisoformat(filter_date_from)
        if filter_date_to:
            date_to = date.fromisoformat(filter_date_to)
        if filter_due_from:
            due_from = date.fromisoformat(filter_due_from)
        if filter_due_to:
            due_to = date.fromisoformat(filter_due_to)
    except (ValueError, TypeError):
        pass

    # Filter invoices within each client group (date/invoice status filters)
    if filter_inv_status or date_from or date_to or due_from or due_to:
        for name, c in client_map.items():
            filtered = c['invoices']
            if filter_inv_status:
                filtered = [inv for inv in filtered if inv.status == filter_inv_status.upper()]
            if date_from:
                filtered = [inv for inv in filtered if inv.date and inv.date >= date_from]
            if date_to:
                filtered = [inv for inv in filtered if inv.date and inv.date <= date_to]
            if due_from:
                filtered = [inv for inv in filtered if inv.due_date and inv.due_date >= due_from]
            if due_to:
                filtered = [inv for inv in filtered if inv.due_date and inv.due_date <= due_to]
            c['invoices'] = filtered
            # Recalculate totals based on filtered invoices
            c['total_invoiced'] = sum(inv.total for inv in filtered)
            c['total_paid'] = sum(inv.amount_paid for inv in filtered if inv.status == 'PAID')
            c['total_outstanding'] = sum(inv.amount_due for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'))
            c['overdue_amount'] = sum(inv.amount_due for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT') and inv.due_date and inv.due_date < today)
            c['invoice_count'] = len(filtered)
            c['outstanding_count'] = len([inv for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT')])
            c['paid_count'] = len([inv for inv in filtered if inv.status == 'PAID'])
            # Recalculate aging buckets
            c['aging'] = {'current': Decimal('0'), '1_30': Decimal('0'), '31_60': Decimal('0'), '61_90': Decimal('0'), '91_120': Decimal('0'), '120_plus': Decimal('0')}
            for inv in filtered:
                if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'):
                    days = inv.days_outstanding
                    if days <= 0:
                        c['aging']['current'] += inv.amount_due
                    elif days <= 30:
                        c['aging']['1_30'] += inv.amount_due
                    elif days <= 60:
                        c['aging']['31_60'] += inv.amount_due
                    elif days <= 90:
                        c['aging']['61_90'] += inv.amount_due
                    elif days <= 120:
                        c['aging']['91_120'] += inv.amount_due
                    else:
                        c['aging']['120_plus'] += inv.amount_due

        # Remove clients with no matching invoices
        clients = sorted([c for c in client_map.values() if c['invoice_count'] > 0], key=lambda x: x['total_outstanding'], reverse=True)
    else:
        clients = sorted(client_map.values(), key=lambda x: x['total_outstanding'], reverse=True)

    # Client-level filters
    if search_q:
        clients = [c for c in clients if search_q.lower() in c['contact_name'].lower()]
    if filter_status == 'outstanding':
        clients = [c for c in clients if c['total_outstanding'] > 0]
    elif filter_status == 'paid':
        clients = [c for c in clients if c['total_outstanding'] == 0 and c['total_paid'] > 0]
    if filter_bucket:
        bucket_key = filter_bucket.replace('-', '_').replace('+', '_plus')
        clients = [c for c in clients if c['aging'].get(bucket_key, 0) > 0]

    # Sorting
    if sort_by == 'name_asc':
        clients.sort(key=lambda x: x['contact_name'].lower())
    elif sort_by == 'name_desc':
        clients.sort(key=lambda x: x['contact_name'].lower(), reverse=True)
    elif sort_by == 'outstanding_asc':
        clients.sort(key=lambda x: x['total_outstanding'])
    elif sort_by == 'outstanding_desc':
        clients.sort(key=lambda x: x['total_outstanding'], reverse=True)
    elif sort_by == 'overdue_desc':
        clients.sort(key=lambda x: x['overdue_amount'], reverse=True)
    elif sort_by == 'oldest':
        clients.sort(key=lambda x: x['oldest_overdue_days'], reverse=True)

    # Grand totals
    grand_invoiced = sum(c['total_invoiced'] for c in clients)
    grand_paid = sum(c['total_paid'] for c in clients)
    grand_outstanding = sum(c['total_outstanding'] for c in clients)
    grand_overdue = sum(c['overdue_amount'] for c in clients)

    # Last sync time
    last_sync = all_invoices.order_by('-synced_at').first()
    last_sync_time = last_sync.synced_at if last_sync else None

    context = {
        'xero_connected': xero_connected,
        'tenant_name': tenant_name,
        'clients': clients,
        'grand_invoiced': grand_invoiced,
        'grand_paid': grand_paid,
        'grand_outstanding': grand_outstanding,
        'grand_overdue': grand_overdue,
        'client_count': len(clients),
        'last_sync_time': last_sync_time,
        'today': today,
        'filter_bucket': filter_bucket,
        'filter_status': filter_status,
        'search_q': search_q,
        'filter_inv_status': filter_inv_status,
        'filter_date_from': filter_date_from,
        'filter_date_to': filter_date_to,
        'filter_due_from': filter_due_from,
        'filter_due_to': filter_due_to,
        'sort_by': sort_by,
    }
    return render(request, 'main/debtors.html', context)


@login_required
def debtors_export_excel(request):
    """Export debtors data to Excel with all current filters applied."""
    if request.user.role not in ('admin', 'super_admin', 'developer'):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Permission denied')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()

    # Get all ACCREC invoices - exclude VOIDED and DELETED
    all_invoices = XeroInvoice.objects.filter(
        invoice_type='ACCREC'
    ).exclude(status__in=['VOIDED', 'DELETED'])

    # Build client-level data (same logic as debtors_page)
    client_map = {}
    for inv in all_invoices:
        name = inv.contact_name or 'Unknown'
        if name not in client_map:
            client_map[name] = {
                'contact_name': name,
                'invoices': [],
                'total_invoiced': Decimal('0'),
                'total_paid': Decimal('0'),
                'total_outstanding': Decimal('0'),
                'overdue_amount': Decimal('0'),
                'oldest_overdue_days': 0,
                'aging': {
                    'current': Decimal('0'), '1_30': Decimal('0'),
                    '31_60': Decimal('0'), '61_90': Decimal('0'),
                    '91_120': Decimal('0'), '120_plus': Decimal('0'),
                },
                'invoice_count': 0, 'outstanding_count': 0, 'paid_count': 0,
            }
        c = client_map[name]
        c['invoices'].append(inv)
        c['invoice_count'] += 1
        c['total_invoiced'] += inv.total
        if inv.status == 'PAID':
            c['total_paid'] += inv.amount_paid
            c['paid_count'] += 1
        elif inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'):
            c['total_outstanding'] += inv.amount_due
            c['outstanding_count'] += 1
            days = inv.days_outstanding
            if days <= 0:
                c['aging']['current'] += inv.amount_due
            elif days <= 30:
                c['aging']['1_30'] += inv.amount_due
            elif days <= 60:
                c['aging']['31_60'] += inv.amount_due
            elif days <= 90:
                c['aging']['61_90'] += inv.amount_due
            elif days <= 120:
                c['aging']['91_120'] += inv.amount_due
            else:
                c['aging']['120_plus'] += inv.amount_due
            if inv.due_date and inv.due_date < today:
                c['overdue_amount'] += inv.amount_due
                c['oldest_overdue_days'] = max(c['oldest_overdue_days'], days)

    for name, c in client_map.items():
        c['invoices'].sort(key=lambda x: x.due_date or date.min)

    # Apply same filters as debtors_page
    filter_bucket = request.GET.get('aging', '')
    filter_status = request.GET.get('status', '')
    search_q = request.GET.get('q', '').strip()
    filter_inv_status = request.GET.get('inv_status', '')
    filter_date_from = request.GET.get('date_from', '')
    filter_date_to = request.GET.get('date_to', '')
    filter_due_from = request.GET.get('due_from', '')
    filter_due_to = request.GET.get('due_to', '')
    sort_by = request.GET.get('sort', '')

    date_from = date_to = due_from = due_to = None
    try:
        if filter_date_from: date_from = date.fromisoformat(filter_date_from)
        if filter_date_to: date_to = date.fromisoformat(filter_date_to)
        if filter_due_from: due_from = date.fromisoformat(filter_due_from)
        if filter_due_to: due_to = date.fromisoformat(filter_due_to)
    except (ValueError, TypeError):
        pass

    if filter_inv_status or date_from or date_to or due_from or due_to:
        for name, c in client_map.items():
            filtered = c['invoices']
            if filter_inv_status:
                filtered = [inv for inv in filtered if inv.status == filter_inv_status.upper()]
            if date_from:
                filtered = [inv for inv in filtered if inv.date and inv.date >= date_from]
            if date_to:
                filtered = [inv for inv in filtered if inv.date and inv.date <= date_to]
            if due_from:
                filtered = [inv for inv in filtered if inv.due_date and inv.due_date >= due_from]
            if due_to:
                filtered = [inv for inv in filtered if inv.due_date and inv.due_date <= due_to]
            c['invoices'] = filtered
            c['total_invoiced'] = sum(inv.total for inv in filtered)
            c['total_paid'] = sum(inv.amount_paid for inv in filtered if inv.status == 'PAID')
            c['total_outstanding'] = sum(inv.amount_due for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'))
            c['overdue_amount'] = sum(inv.amount_due for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT') and inv.due_date and inv.due_date < today)
            c['invoice_count'] = len(filtered)
            c['outstanding_count'] = len([inv for inv in filtered if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT')])
            c['paid_count'] = len([inv for inv in filtered if inv.status == 'PAID'])
            c['aging'] = {'current': Decimal('0'), '1_30': Decimal('0'), '31_60': Decimal('0'), '61_90': Decimal('0'), '91_120': Decimal('0'), '120_plus': Decimal('0')}
            for inv in filtered:
                if inv.status in ('AUTHORISED', 'SUBMITTED', 'DRAFT'):
                    days = inv.days_outstanding
                    if days <= 0: c['aging']['current'] += inv.amount_due
                    elif days <= 30: c['aging']['1_30'] += inv.amount_due
                    elif days <= 60: c['aging']['31_60'] += inv.amount_due
                    elif days <= 90: c['aging']['61_90'] += inv.amount_due
                    elif days <= 120: c['aging']['91_120'] += inv.amount_due
                    else: c['aging']['120_plus'] += inv.amount_due
        clients = sorted([c for c in client_map.values() if c['invoice_count'] > 0], key=lambda x: x['total_outstanding'], reverse=True)
    else:
        clients = sorted(client_map.values(), key=lambda x: x['total_outstanding'], reverse=True)

    if search_q:
        clients = [c for c in clients if search_q.lower() in c['contact_name'].lower()]
    if filter_status == 'outstanding':
        clients = [c for c in clients if c['total_outstanding'] > 0]
    elif filter_status == 'paid':
        clients = [c for c in clients if c['total_outstanding'] == 0 and c['total_paid'] > 0]
    if filter_bucket:
        bucket_key = filter_bucket.replace('-', '_').replace('+', '_plus')
        clients = [c for c in clients if c['aging'].get(bucket_key, 0) > 0]

    if sort_by == 'name_asc': clients.sort(key=lambda x: x['contact_name'].lower())
    elif sort_by == 'name_desc': clients.sort(key=lambda x: x['contact_name'].lower(), reverse=True)
    elif sort_by == 'outstanding_asc': clients.sort(key=lambda x: x['total_outstanding'])
    elif sort_by == 'outstanding_desc': clients.sort(key=lambda x: x['total_outstanding'], reverse=True)
    elif sort_by == 'overdue_desc': clients.sort(key=lambda x: x['overdue_amount'], reverse=True)
    elif sort_by == 'oldest': clients.sort(key=lambda x: x['oldest_overdue_days'], reverse=True)

    # Build Excel workbook
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = 'Debtors Summary'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='007890', end_color='007890', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    money_fmt = '#,##0.00'

    summary_headers = ['Client', 'Invoices', 'Outstanding', 'Total Invoiced', 'Total Paid',
                        'Overdue', 'Current', '1-30 Days', '31-60 Days', '61-90 Days',
                        '91-120 Days', '120+ Days']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, client in enumerate(clients, 2):
        vals = [
            client['contact_name'], client['outstanding_count'],
            float(client['total_outstanding']), float(client['total_invoiced']),
            float(client['total_paid']), float(client['overdue_amount']),
            float(client['aging']['current']), float(client['aging']['1_30']),
            float(client['aging']['31_60']), float(client['aging']['61_90']),
            float(client['aging']['91_120']), float(client['aging']['120_plus']),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 3:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right')

    # Totals row
    total_row = len(clients) + 2
    ws.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=sum(c['outstanding_count'] for c in clients)).font = Font(bold=True)
    for col_idx, key in enumerate(['total_outstanding', 'total_invoiced', 'total_paid', 'overdue_amount'], 3):
        cell = ws.cell(row=total_row, column=col_idx, value=float(sum(c[key] for c in clients)))
        cell.font = Font(bold=True)
        cell.number_format = money_fmt
        cell.border = thin_border
    for col_idx, key in enumerate(['current', '1_30', '31_60', '61_90', '91_120', '120_plus'], 7):
        cell = ws.cell(row=total_row, column=col_idx, value=float(sum(c['aging'][key] for c in clients)))
        cell.font = Font(bold=True)
        cell.number_format = money_fmt
        cell.border = thin_border

    # Auto-width columns
    for col_idx in range(1, len(summary_headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, total_row + 1))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)

    # --- Sheet 2: Invoice Detail ---
    ws2 = wb.create_sheet('Invoice Detail')
    detail_headers = ['Client', 'Invoice #', 'Reference', 'Status', 'Invoice Date',
                      'Due Date', 'Total', 'Amount Paid', 'Amount Due', 'Days Outstanding']
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    detail_row = 2
    for client in clients:
        for inv in client['invoices']:
            vals = [
                client['contact_name'], inv.invoice_number or '', inv.reference or '',
                inv.status, inv.date, inv.due_date,
                float(inv.total), float(inv.amount_paid), float(inv.amount_due),
                inv.days_outstanding,
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws2.cell(row=detail_row, column=col_idx, value=val)
                cell.border = thin_border
                if col_idx in (5, 6) and val:
                    cell.number_format = 'YYYY-MM-DD'
                if col_idx in (7, 8, 9):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal='right')
            detail_row += 1

    for col_idx in range(1, len(detail_headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(ws2.cell(row=r, column=col_idx).value or '')) for r in range(1, min(detail_row, 50)))
        ws2.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)

    # Write to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Debtors_Report_{today.strftime("%Y%m%d")}.xlsx"'
    return response
